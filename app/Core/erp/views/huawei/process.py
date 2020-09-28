from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment

from openpyxl.worksheet.worksheet import Worksheet

g = globals()


def write_shell_workbook(header: list, header_parameter: list, sheet_name_new: Worksheet):
    i = 0
    for rows in header:
        sheet_name_new[rows] = header_parameter[i]
        sheet_name_new[rows].font = Font(name='Arial', size=10, color="000000", bold=True, vertAlign=None)
        sheet_name_new[rows].alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        i = i + 1


def convert_string_to_int(value: str):
    try:
        response = int(value)
        return True
    except KeyError:
        return False


def get_sheet_name_lte_df(sheet_name_new: Worksheet, sheet_names_df: list):
    list_get_sheet: list = []
    for i in range(2, sheet_name_new.max_row):
        mo = ""
        mo = str(sheet_name_new.cell(row=i, column=2).value).upper().strip()
        if len(mo) > 0 and mo != 'None':
            if mo in sheet_names_df:
                if mo in list_get_sheet:
                    k = 0
                else:
                    list_get_sheet.append(mo)
    return list_get_sheet


def get_sheet_name_lte_df_other(sheet_name_new: Worksheet, sheet_names_df: list, sheet_name_group: list = []):
    list_get_sheet: list = []
    for i in range(2, sheet_name_new.max_row):
        mo = ""
        mo = str(sheet_name_new.cell(row=i, column=2).value).upper().strip()
        if len(mo) > 0 and mo != 'None':
            if mo not in sheet_names_df and mo not in sheet_name_group:
                if mo in list_get_sheet:
                    k = 0
                else:
                    list_get_sheet.append(mo)
    return list_get_sheet


def get_sheet_name_lte_distinct_df(sheet_name_new: Worksheet, sheet_names_df: list):
    list_get_sheet: list = []
    for i in range(2, sheet_name_new.max_row):
        mo = str(sheet_name_new.cell(row=i, column=2).value).upper().strip()
        if len(mo) > 0 and mo != 'None':
            if mo in list_get_sheet:
                k = 0
            else:
                if mo not in sheet_names_df:
                    list_get_sheet.append(mo)

    return list_get_sheet


def get_sheet_name_case_1_group_id(sheet_name_new: Worksheet, sheet_names_df: list):
    list_get_sheet: list = []
    for i in range(2, sheet_name_new.max_row):
        mo = str(sheet_name_new.cell(row=i, column=2).value).upper().strip()
        ParameterName = str(sheet_name_new.cell(row=i, column=4).value).upper().strip()
        ParameterValue = str(sheet_name_new.cell(row=i, column=6).value).upper().strip()
        if len(mo) > 0 and mo != 'None':
            if len(ParameterName) > 8:
                ParameterNamev = ParameterName[-8:]
                if ParameterNamev == "group ID".upper() and ParameterValue.find('|') == -1 and mo != "X2":
                    if (mo in list_get_sheet) is False and (mo in sheet_names_df) is False:
                        list_get_sheet.append(mo)

    return list_get_sheet


def get_parameter_by_mo(sheet_name_new: Worksheet, mo_name: str):
    list_parameter_mo = {}
    Min_row = 2
    Max_row = sheet_name_new.max_row + 1
    for i in range(Min_row, Max_row):
        VP_MO_Shell = str(sheet_name_new.cell(row=i, column=2).value).upper().strip()
        if VP_MO_Shell == mo_name.upper():
            VP_ParameterID = str(sheet_name_new.cell(row=i, column=3).value).strip()
            VP_ParameterName = str(sheet_name_new.cell(row=i, column=4).value).strip().title()

            VP_FeatureValue = str(sheet_name_new.cell(row=i, column=6).value).strip()
            if VP_MO_Shell == "UTRANNCELL":
                if VP_ParameterName == "Neighbour Cell Name":
                    VP_FeatureValue = "<Cellname>"
                elif VP_ParameterName == "Local Cell Name":
                    VP_FeatureValue = "<LocalCellName>"
                elif VP_ParameterName == "Cell Measure Priority" and str(
                        VP_FeatureValue).upper() == "Low Priority".upper():
                    VP_FeatureValue = "LOW_PRIORITY"
            list_parameter_mo[str(VP_ParameterName).title()] = str(VP_FeatureValue)
    return list_parameter_mo


def compare_two_string(a, b):
    s = ""
    for i in range(len(a)):
        if a[i] < b[i]:
            s += str(a[i])
        else:
            s += str(b[i])
    return s


def get_df4g_row_site(sheet_name_new: Worksheet):
    # Sheet  =>  DF 4G
    list_row_df4g = {}
    row = {}
    Min_row = 2
    Max_row = sheet_name_new.max_row + 1
    row_item = 0
    for i in range(Min_row, Max_row):
        value = str(sheet_name_new.cell(row=i, column=5).value).upper().lstrip().rstrip()
        row = {}
        if len(value) > 0 and value != 'None' and value != "NONE":
            row_item = row_item + 1
            row["NE_Name_df4g"] = str(sheet_name_new.cell(row=i, column=5).value).upper().strip()
            row["eNodeB_Name_df4g"] = str(sheet_name_new.cell(row=i, column=6).value).upper().strip()
            row["CellId_df4g"] = str(sheet_name_new.cell(row=i, column=19).value).upper().strip()
            list_row_df4g["site_" + str(row_item)] = row
    return list_row_df4g


def group_case_1_get_parameter_by_id_group(list_parameter: dict, cell_id_group_validate: int):
    list_case1 = {}
    General_list_case1 = {}
    VP_Group = ""
    VP_Group_Val = False
    list_not_group_id = {}
    flat = True
    item_Group = 0
    item_g = 0
    parameter_contador = 0
    for parameter_name, value in list_parameter.items():
        parameter_contador = parameter_contador + 1
        parameter_name = str(parameter_name).strip().title()
        parameter_name = str(parameter_name[:-len(str(parameter_contador))]).strip().title()
        if VP_Group_Val is False:
            if len(parameter_name) > 8:
                ParameterNamev = parameter_name[-8:].upper()
                if ParameterNamev == "group ID".upper():
                    VP_Group = parameter_name
                    VP_Group_Val = True
                else:
                    VP_Group_Val = False
            else:
                VP_Group_Val = False

            if VP_Group_Val is False:
                # tengo que validar si existe el ["Local Cell ID"]
                list_not_group_id[parameter_name] = value

        if VP_Group_Val is True:
            if VP_Group == parameter_name:
                item_Group = item_Group + 1
                if item_Group == 1:
                    list_case1 = {}
                    list_case1[parameter_name] = value
                else:
                    item_g = item_g + 1
                    item_Group = 0
                    if len(list_not_group_id) > 0:
                        distinct_item = dict(list_case1, **list_not_group_id)
                        General_list_case1["key" + str(item_g)] = distinct_item
                    else:
                        if cell_id_group_validate > 0:
                            list_case1["Local Cell ID"] = 0

                        General_list_case1["key" + str(item_g)] = list_case1
                        list_case1 = {}
                        list_case1[parameter_name] = value
                        VP_Group_Val = True

            else:
                item_Group = item_Group + 1
                list_case1[parameter_name] = value

    if len(list_parameter.items()) == parameter_contador:
        item_g = item_g + 1
        if len(list_not_group_id) > 0:
            distinct_item = dict(list_case1, **list_not_group_id)
            General_list_case1["key" + str(item_g)] = distinct_item
        else:
            if cell_id_group_validate > 0:
                list_case1["Local Cell ID"] = 0
            General_list_case1["key" + str(item_g)] = list_case1
    return General_list_case1


def group_case_1_get_get_position_by_id_group(list_parameter: dict, sheet_name_site: Worksheet):
    list_new_parameter: list = []
    list_new_position = {}
    post = 0
    for list_parameter_key, list_parameter_value in list_parameter.items():

        list_parameter_input = dict(list_parameter_value)
        list_new_parameter = []
        for parameter_key, parameter_value in list_parameter_input.items():
            list_new_parameter.append(str(parameter_key).upper().strip())
        list_position_header_site = {}
        list_position_header_item = 0
        CS1_INDEX_FIND_SITE = 0
        CS1_INDEX_SITE = 0
        CS1_COLUM_VALUE = ""
        for parameter_name in list_new_parameter:
            list_position_header_item += 1
            for u in range(1, (sheet_name_site.max_column + 1)):
                CS1_COLUM_VALUE = str(
                    sheet_name_site.cell(row=2, column=u).value).upper().lstrip().rstrip()
                CS1_COLUM_VALUE = CS1_COLUM_VALUE.replace("*", "")
                CS1_INDEX_FIND_SITE = CS1_COLUM_VALUE.find(parameter_name.upper())
                if CS1_INDEX_FIND_SITE == 0:
                    # validar si existe '('
                    ValueString = ""
                    indexletra = CS1_COLUM_VALUE.find('(')
                    if indexletra > 0:
                        ValueString = CS1_COLUM_VALUE[:indexletra].strip().rstrip()
                    else:
                        ValueString = CS1_COLUM_VALUE

                    mo = compare_two_string(parameter_name.upper(),
                                            CS1_COLUM_VALUE).upper().lstrip().rstrip()
                    if mo == parameter_name.upper() and len(ValueString) == len(mo):
                        CS1_INDEX_SITE = u
                        break
            if CS1_INDEX_SITE > 0:
                list_position_header_site['Col_' + str(list_position_header_item)] = CS1_INDEX_SITE
            else:
                list_position_header_site['Col_' + str(list_position_header_item)] = -1
        post = post + 1
        list_new_position["key_" + str(post)] = list_position_header_site

    return list_new_position


def get_position_other(list_parameter: dict, sheet_name_site: Worksheet):
    list_new_parameter: list = []
    list_new_position = {}
    post = 0
    parameter_contador = 0
    list_position_header_site = {}
    list_position_header_item = 0
    for list_parameter_key, list_parameter_value in list_parameter.items():

        CS1_INDEX_FIND_SITE = 0
        CS1_INDEX_SITE = 0
        CS1_COLUM_VALUE = ""
        parameter_contador = parameter_contador + 1
        parameter_name = str(list_parameter_key).strip().title()
        parameter_name = str(parameter_name[:-len(str(parameter_contador))]).strip().title()
        list_position_header_item += 1
        for u in range(1, (sheet_name_site.max_column + 1)):
            CS1_COLUM_VALUE = str(
                sheet_name_site.cell(row=2, column=u).value).upper().lstrip().rstrip()
            CS1_COLUM_VALUE = CS1_COLUM_VALUE.replace("*", "")
            CS1_INDEX_FIND_SITE = CS1_COLUM_VALUE.find(parameter_name.upper())
            if CS1_INDEX_FIND_SITE == 0:
                # validar si existe '('
                ValueString = ""
                indexletra = CS1_COLUM_VALUE.find('(')
                if indexletra > 0:
                    ValueString = CS1_COLUM_VALUE[:indexletra].strip().rstrip()
                else:
                    ValueString = CS1_COLUM_VALUE

                mo = compare_two_string(parameter_name.upper(),
                                        CS1_COLUM_VALUE).upper().lstrip().rstrip()
                if mo == parameter_name.upper() and len(ValueString) == len(mo):
                    CS1_INDEX_SITE = u
                    break
        if CS1_INDEX_SITE > 0:
            list_position_header_site['Col_' + str(list_position_header_item)] = CS1_INDEX_SITE
        else:
            list_position_header_site['Col_' + str(list_position_header_item)] = -1
    post = post + 1
    list_new_position["key_" + str(post)] = list_position_header_site

    return list_new_position


def get_new_parameter_by_mo(list_parameter_df: list, list_parameter_lte: dict):
    list_lte: list = []
    value_new = ""
    for key in list_parameter_lte.keys():
        value_new = str(key).title()
        list_lte.append(value_new.strip())
    lisparameterconvert = [element.title().strip() for element in list_parameter_df]
    response = list(set(list_lte).difference(lisparameterconvert))
    # for key in response:
    #    list_parameter_df.append(str(key).title())
    return response


def add_new_parameter_by_mo(list_parameter_df: list, list_new_parameter: list):
    for key in list_new_parameter:
        list_parameter_df.append(str(key).title())
    return list_parameter_df


# obtiene los valores del DF
def add_new_parameter_correct_value(list_value_output: list, list_parameter_lte: dict, list_new_parameter: list):
    for new_parameter in list_new_parameter:
        for lte_parameter, lte_parameter_value in list_parameter_lte.items():
            lte_parameter_name = str(lte_parameter)
            if lte_parameter_name == new_parameter:
                list_value_output.append(lte_parameter_value.upper().strip())

    return list_value_output


# obtiene los valores del Site
def add_new_parameter_current_value(list_value_output: list, sheet_name_site: Worksheet,
                                    list_new_parameter_position: dict, position_row: int):
    for key, position_column in list_new_parameter_position.items():
        if int(position_column) > 0:
            value = str(sheet_name_site.cell(row=position_row, column=position_column).value).upper().strip()
            list_value_output.append(value)
        else:
            list_value_output.append("COLUMN_NOT_FOUND")

    return list_value_output


def get_position_new_parameter(list_new_parameter: list, sheet_name_site: Worksheet):
    list_position_header_site = {}
    list_position_header_item = 0
    CS1_INDEX_FIND_SITE = 0
    CS1_INDEX_SITE = 0
    CS1_COLUM_VALUE = ""
    for parameter_name in list_new_parameter:
        list_position_header_item += 1
        for u in range(1, (sheet_name_site.max_column + 1)):
            CS1_COLUM_VALUE = str(
                sheet_name_site.cell(row=2, column=u).value).upper().lstrip().rstrip()
            CS1_COLUM_VALUE = CS1_COLUM_VALUE.replace("*", "")
            CS1_INDEX_FIND_SITE = CS1_COLUM_VALUE.find(parameter_name.upper())
            if CS1_INDEX_FIND_SITE == 0:
                # validar si existe '('
                ValueString = ""
                indexletra = CS1_COLUM_VALUE.find('(')
                if indexletra > 0:
                    ValueString = CS1_COLUM_VALUE[:indexletra].strip().rstrip()
                else:
                    ValueString = CS1_COLUM_VALUE

                mo = compare_two_string(parameter_name.upper(),
                                        CS1_COLUM_VALUE).upper().lstrip().rstrip()
                if mo == parameter_name.upper() and len(ValueString) == len(mo):
                    CS1_INDEX_SITE = u
                    break
        if CS1_INDEX_SITE > 0:
            list_position_header_site['Col_' + str(list_position_header_item)] = CS1_INDEX_SITE
        else:
            list_position_header_site['Col_' + str(list_position_header_item)] = -1

    return list_position_header_site


class Validate_Huawei:
    excel_document_LTE: Workbook
    excel_document_CONF_SITE1: Workbook
    excel_document_CONF_SITE2: Workbook
    excel_document_CONF_SITE3: Workbook
    execel_document_Data_prameter3g_cell: Workbook
    execel_document_Data_prameter3gRadNetw: Workbook
    excel_document_DF: Workbook
    wb: Workbook
    list_sheet_namesDF: list
    # Parametros Generales ->
    excel_NameSheet_LTE = 'Baseline_Unique'
    excel_NameWorkbook_LTE = 'NTA_LTE_Ran_Sharing_IPT_Minimacro_V4_20200805'
    Excel_sheet_LTE: object
    dest_filename: str
    ws1: Worksheet
    ws2: Worksheet
    ws3: Worksheet

    # Configuracion SITE
    excel_NameWorkbook_CONF_SITE1 = 'ConfigurationData_Cell'
    excel_NameWorkbook_CONF_SITE2 = 'RnpData_BTS3900'
    excel_NameWorkbook_CONF_SITE3 = 'ConfigurationData_eNodeB'

    INDEX_Excel_LTE_Write_Sheet1: int = 3

    def __init__(self, url_lte, url_site1, url_site2, url_site3, url_df, url_3g_cell, url_3g_rad_net, name_file_lte):
        self.excel_document_LTE = openpyxl.load_workbook(filename=url_lte)
        # Directorios de archivo  de configuración del Site
        self.excel_document_CONF_SITE1 = openpyxl.load_workbook(filename=url_site1)
        self.excel_document_CONF_SITE2 = openpyxl.load_workbook(filename=url_site2)
        self.excel_document_CONF_SITE3 = openpyxl.load_workbook(filename=url_site3)
        self.execel_document_Data_prameter3g_cell = openpyxl.load_workbook(filename=url_3g_cell)
        self.execel_document_Data_prameter3gRadNetw = openpyxl.load_workbook(filename=url_3g_rad_net)
        # Directorios de archivo  de DF
        self.excel_document_DF = openpyxl.load_workbook(filename=url_df)
        self.list_sheet_namesDF = self.excel_document_DF.sheetnames
        self.excel_NameWorkbook_LTE = name_file_lte
        self.excel_NameSheet_LTE = 'Baseline_Unique'
        self.wb = Workbook()

        self.ws1: Worksheet = self.wb.active
        self.ws1.title = "Workbook Validate"
        self.ws2 = self.wb.create_sheet(title="Result")
        self.ws3 = self.wb.create_sheet(title="ResultDetail")

        header_ws1 = ["B3", "C3", "D3", "E3", "F3", "G3"]
        header_parameter_ws1 = ["FILE TYPE", "NAME FILE", "NAME SHEET", "REQUIRED", "IT IS VALID", "COMMENT"]

        header_ws2 = ["B3", "C3", "D3", "E3", "F3", "G3", "H3"]
        header_parameter_ws2 = ["CELL-ID", "MO", "Parameter Name", "Value correct", "Value current", "STATUS",
                                "COMMENT"]

        header_ws3 = ["B3", "C3", "D3", "E3", "F3", "G3", "H3", "I3", "J3"]
        header_parameter_ws3 = ["ResultDetailID", "ResultID", "RowId", "ParameterID", "Parameter Name", "Value CORRECT",
                                "Value Correct", "STATUS", "COMMENT"]

        write_shell_workbook(header=header_ws1, header_parameter=header_parameter_ws1, sheet_name_new=self.ws1)
        write_shell_workbook(header=header_ws2, header_parameter=header_parameter_ws2, sheet_name_new=self.ws2)
        write_shell_workbook(header=header_ws3, header_parameter=header_parameter_ws3, sheet_name_new=self.ws3)
        self.INDEX_Excel_LTE_Write_Sheet1 = self.INDEX_Excel_LTE_Write_Sheet1 + 1

    def save_workbook(self,dest_filename:str):
        self.wb.save(filename=dest_filename)

    def validate_general_huawei(self, dest_filename: str):

        list_sheet_namesDF = self.excel_document_DF.sheetnames

        ValidateExistenceSheet_LTE = self.validate_exists_sheet_excel_by_workbook(self.INDEX_Excel_LTE_Write_Sheet1,
                                                                                  'LTE',
                                                                                  self.excel_NameSheet_LTE,
                                                                                  self.excel_NameWorkbook_LTE,
                                                                                  self.excel_document_LTE, True, False)
        if ValidateExistenceSheet_LTE:
            Excel_sheet_LTE = self.excel_document_LTE[self.excel_NameSheet_LTE]

            index_result_1 = 4  # Hoja Principal  (validacion de que exista la hoja)
            index_result_2 = 4  # Hoja Principal  (Panel de resultados)
            index_result_3 = 0  # Hoja Principal  (Panel de resultados detalle)

            list_output = {}  # recibe los index de las hojas:

            listNameDistinctDF = get_sheet_name_lte_df(sheet_name_new=Excel_sheet_LTE,
                                                       sheet_names_df=list_sheet_namesDF)
            # Validate Config 1
            list_output = self.validate_sheet_df(excel_sheet_lte=Excel_sheet_LTE,
                                                 workbook_site1=self.excel_document_CONF_SITE1,
                                                 workbook_site2=self.excel_document_CONF_SITE2,
                                                 list_name_df=listNameDistinctDF, index_result_1=index_result_1,
                                                 index_result_2=index_result_2)
            index_result_1 = list_output["index_result_1"]
            index_result_2 = list_output["index_result_2"]

            sheet_name_case_group = get_sheet_name_case_1_group_id(sheet_name_new=Excel_sheet_LTE,
                                                                   sheet_names_df=listNameDistinctDF)

            list_output = self.validate_sheet_group_case_1(excel_sheet_lte=Excel_sheet_LTE,
                                                           workbook_site1=self.excel_document_CONF_SITE1,
                                                           workbook_site2=self.excel_document_CONF_SITE2,
                                                           workbook_site3=self.excel_document_CONF_SITE3,
                                                           list_name_df=sheet_name_case_group,
                                                           index_result_1=index_result_1,
                                                           index_result_2=index_result_2)
            list_sheet_name_lte_other = get_sheet_name_lte_df_other(sheet_name_new=Excel_sheet_LTE,
                                                                    sheet_names_df=listNameDistinctDF,
                                                                    sheet_name_group=sheet_name_case_group)

            index_result_1 = list_output["index_result_1"]
            index_result_2 = list_output["index_result_2"]

            list_output = self.validate_sheet_others(excel_sheet_lte=Excel_sheet_LTE,
                                                     workbook_site1=self.excel_document_CONF_SITE1,
                                                     workbook_site2=self.excel_document_CONF_SITE2,
                                                     workbook_site3=self.excel_document_CONF_SITE3,
                                                     list_name_df=list_sheet_name_lte_other,
                                                     index_result_1=index_result_1,
                                                     index_result_2=index_result_2)

            index_result_1 = list_output["index_result_1"]
            index_result_2 = list_output["index_result_2"]

            Excel_sheet_DF = self.excel_document_DF["UCELLNFREQPRIOINFO"]
            self.INDEX_Excel_LTE_Write_Sheet1 = self.INDEX_Excel_LTE_Write_Sheet1 + 1
            ValidateExistenceSheet_Data = self.validate_exists_sheet_excel_by_workbook(
                self.INDEX_Excel_LTE_Write_Sheet1,
                '3G Cell Parameters Data Template',
                "CELLNFREQPRIOINFO",
                "3G Cell Parameters Data Template",
                self.execel_document_Data_prameter3g_cell, True,
                False)
            if ValidateExistenceSheet_Data:
                Excel_sheet_site = self.execel_document_Data_prameter3g_cell["CELLNFREQPRIOINFO"]
                index_result_2 = self.validate_sheet_UCELLNFREQPRIOINFO(p_excel_sheet_df=Excel_sheet_DF,
                                                                        excel_sheet_site=Excel_sheet_site,
                                                                        index_result_2=index_result_2)
            self.INDEX_Excel_LTE_Write_Sheet1 = self.INDEX_Excel_LTE_Write_Sheet1 + 1
            Excel_sheet_DF = self.excel_document_DF["ULTENCELL"]
            ValidateExistenceSheet_Data = self.validate_exists_sheet_excel_by_workbook(
                self.INDEX_Excel_LTE_Write_Sheet1,
                '3G Radio Network Planning Data Template',
                "LTENCELL",
                "3G Radio Network Planning Data Template",
                self.execel_document_Data_prameter3gRadNetw, True,
                False)
            if ValidateExistenceSheet_Data:
                Excel_sheet_site = self.execel_document_Data_prameter3gRadNetw["LTENCELL"]
                index_result_2 = self.validate_sheet_ULTENCELL(p_excel_sheet_df=Excel_sheet_DF,
                                                               excel_sheet_site=Excel_sheet_site,
                                                               index_result_2=index_result_2)
            return list_output

    def validate_exists_sheet_excel_by_workbook(self, index_excel_lte_write: int, filetype: str, excel_name_sheet: str,
                                                excel_name_workbook: str, excel_document: Workbook, mo_new: bool,
                                                process_new: bool):
        excel_sheet_generic = object
        response: bool = False
        try:
            excel_sheet_generic = excel_document[excel_name_sheet]
            if mo_new and process_new is False:
                self.ws1['B' + str(index_excel_lte_write)] = filetype
                self.ws1['C' + str(index_excel_lte_write)] = excel_name_workbook
                self.ws1['D' + str(index_excel_lte_write)] = excel_name_sheet
                self.ws1['E' + str(index_excel_lte_write)] = "YES"
                self.ws1['F' + str(index_excel_lte_write)] = "SI"
                self.ws1['G' + str(index_excel_lte_write)] = "Configuration sheet : " + excel_name_sheet
            return True
        except KeyError:
            if mo_new and process_new is False:
                self.ws1['B' + str(index_excel_lte_write)] = filetype
                self.ws1['C' + str(index_excel_lte_write)] = excel_name_workbook
                self.ws1['D' + str(index_excel_lte_write)] = excel_name_sheet
                self.ws1['E' + str(index_excel_lte_write)] = "YES"
                self.ws1['F' + str(index_excel_lte_write)] = "NOT"
                self.ws1['G' + str(index_excel_lte_write)] = "Configuration sheet not found: " + excel_name_sheet
            return False

    def get_parameter_mo_by_group(self, sheet_name_new: Worksheet, mo_name: str):
        list_parameter_mo = {}
        Min_row = 2
        Max_row = sheet_name_new.max_row + 1
        item_count = 0
        for i in range(Min_row, Max_row):
            VP_MO_Shell = str(sheet_name_new.cell(row=i, column=2).value).upper().strip()
            if VP_MO_Shell == mo_name.upper().strip():

                VP_ParameterName = str(sheet_name_new.cell(row=i, column=4).value).strip().title()
                VP_FeatureValue = str(sheet_name_new.cell(row=i, column=6).value).strip()
                item_count = item_count + 1
                if VP_MO_Shell == "SCTPTEMPLATE" and VP_ParameterName == "Heart - beat Interval".title():
                    VP_ParameterName = "Heart-Beat Interval"
                if VP_MO_Shell == "GLOBALPROCSWITCH" and VP_ParameterName == "Ho And E-Rab Conflict Processing Strategy".title():
                    VP_FeatureValue = VP_FeatureValue.replace(":On", "").replace(":ON", "").replace(":Off", "").replace(
                        ":OFF", "")
                if VP_MO_Shell == "CELLALGOSWITCH" and VP_ParameterName == "Amr Control Algorithm Switch":
                    VP_ParameterName = "Rate Control Algorithm Switch"
                if VP_MO_Shell == "ANR" and VP_ParameterName in ["Statistic Cycle For Delete Nrt",
                                                                 "Statistic Number For Delete Nrt"]:
                    VP_FeatureValue = VP_FeatureValue.replace('(Non-Lima)', '').strip()
                if VP_MO_Shell == "PDSCHCFG" and VP_ParameterName == "Reference Signal Power":
                    sheet_name_new = self.excel_document_DF["DF 4G"]
                    VP_FeatureValue = str(sheet_name_new.cell(row=2, column=30).value).upper().lstrip().rstrip()
                if VP_MO_Shell in ["SCTPHOST", "USERPLANEHOST"] and VP_ParameterName in ["First Local Ip Address",
                                                                                         "Local Ip Address"]:
                    item_count = item_count - 1
                else:
                    if sheet_name_new.cell(row=i, column=4).font.strike:
                        item_count = item_count - 1
                    else:
                        list_parameter_mo[str(VP_ParameterName).title() + str(item_count)] = str(VP_FeatureValue)

        return list_parameter_mo

    # Permite escribir los resultados en la Hoja "Result"
    def write_ws2_result_mo_parameter(self, identificated_row: str, mo: str, list_parameter: list, index: int):
        for parameter in list_parameter:
            self.ws2["B" + str(index)] = identificated_row
            self.ws2["C" + str(index)] = mo
            self.ws2["D" + str(index)] = parameter
            index = index + 1

    def write_ws2_result_value_correct_current(self, list_parameter_mo: dict, list_parameter_df: list,
                                               list_value_correct: list, list_value_current: list,
                                               index: int, flat_fail: bool = False, message: str = ""):
        value_current: str = "",
        value_parameter_mo = ""
        # message_list: str = ""
        index_value = 0
        for parameter_name in list_parameter_df:
            value_parameter_mo = "NOT_PARAMETER_CONF"
            if str(parameter_name).title() in list_parameter_mo.keys():
                value_parameter_mo = str(list_parameter_mo[parameter_name]).upper() \
                    .strip().replace("'", '').replace("NULL", 'NONE')

            value_correct = str(list_value_correct[index_value]).upper().strip().replace("NULL", 'NONE')
            if flat_fail is False:
                value_current = str(list_value_current[index_value]).upper().strip().replace("NULL", 'NONE')
            if flat_fail is False:
                if (value_parameter_mo.find(">") >= 0 and value_parameter_mo.find(
                        "<") >= 0) or value_parameter_mo == 'NOT_PARAMETER_CONF':

                    if value_correct == value_current:
                        self.ws2["E" + str(index)] = value_correct
                        self.ws2["F" + str(index)] = value_current
                        self.ws2["G" + str(index)] = "Correct"
                        self.ws2["H" + str(index)] = "Correct Value Configurate"
                    else:
                        self.ws2["E" + str(index)] = value_correct
                        self.ws2["F" + str(index)] = value_current
                        self.ws2["G" + str(index)] = "Fail"
                        self.ws2["H" + str(index)] = "Fail Value Configurate"

                else:
                    if parameter_name in ["Mobile Country Code", "Mobile Network Code"]:
                        if value_correct == value_current:
                            self.ws2["E" + str(index)] = value_correct
                            self.ws2["F" + str(index)] = value_current
                            self.ws2["G" + str(index)] = "Correct"
                            self.ws2["H" + str(index)] = "Correct Value Configurate"
                        else:
                            self.ws2["E" + str(index)] = value_correct
                            self.ws2["F" + str(index)] = value_current
                            self.ws2["G" + str(index)] = "Fail"
                            self.ws2["H" + str(index)] = "Fail Value Configurate"
                    else:
                        if value_parameter_mo.upper() == "L700_Cell".upper() \
                                or value_parameter_mo.upper() == "LoCell_700 | LoCell_AWS".upper() or value_parameter_mo.upper() == "LTE_700 Cell | LTE_AWS Cell".upper():
                            if convert_string_to_int(value_current):
                                if 100 >= int(value_current) >= 91:
                                    self.ws2["E" + str(index)] = value_parameter_mo
                                    self.ws2["F" + str(index)] = value_current
                                    self.ws2["G" + str(index)] = "Correct"
                                    self.ws2["H" + str(index)] = "Correct Value Configurate - Is L700_Cell"
                                elif 103 >= int(value_current) >= 101:
                                    self.ws2["E" + str(index)] = value_parameter_mo
                                    self.ws2["F" + str(index)] = value_current
                                    self.ws2["G" + str(index)] = "Correct"
                                    self.ws2["H" + str(index)] = "Correct Value Configurate - Is LTE_AWS"

                                else:
                                    self.ws2["E" + str(index)] = value_parameter_mo
                                    self.ws2["F" + str(index)] = value_current
                                    self.ws2["G" + str(index)] = "Correct"
                                    self.ws2["H" + str(index)] = "Fail Value Configurate - Not is L700_Cell"
                        else:

                            if value_correct == value_current:
                                self.ws2["E" + str(index)] = value_correct
                                self.ws2["F" + str(index)] = value_current
                                self.ws2["G" + str(index)] = "Correct"
                                self.ws2["H" + str(index)] = "Correct Value Configurate"
                                # Se compara el parámetro de LTE  y el parámetro configurado  en el DF
                            # elif value_parameter_mo == value_current and value_parameter_mo != value_correct:
                            #   ws2["E" + str(index)] = value_parameter_mo
                            #    ws2["F" + str(index)] = value_current
                            #    ws2["G" + str(index)] = "Fail"
                            #    ws2["H" + str(index)] = "Fail Value Configurate - Value DF Incorrect: " + value_correct
                            else:
                                self.ws2["E" + str(index)] = value_correct
                                self.ws2["F" + str(index)] = value_current
                                self.ws2["G" + str(index)] = "Fail"
                                self.ws2["H" + str(index)] = "Fail Value Configurate"
            else:
                self.ws2["E" + str(index)] = value_correct
                self.ws2["F" + str(index)] = "-"
                self.ws2["G" + str(index)] = "Fail"
                self.ws2["H" + str(index)] = message
            index = index + 1
            index_value = index_value + 1
        return index

    def write_ws2_group_case_1_value_correct_current(self, cell_id: int, list_parameter_mo: dict,
                                                     list_parameter_site: dict,
                                                     index: int, mo_name: str):
        value_current: str = "",

        index_value = 0
        list_parameter_mo_val: dict = {}
        list_parameter_site_val: dict = {}
        first_Item = ""
        item_count = 0
        for list_mo_key, list_mo_value in list_parameter_mo.items():
            index_value = index_value + 1
            list_parameter_mo_val = dict(list_mo_value)
            list_parameter_site_val = list_parameter_site["key_" + str(index_value)]
            item_count = 0
            for list_parameter_mo_val_key, list_parameter_mo_val_value in list_parameter_site_val.items():
                item_count = item_count + 1

                list_parameter_mo_val_value = str(list_parameter_mo_val_value).upper().strip().replace("'", '').replace(
                    "NULL", 'NONE')
                value_site = str(list_parameter_site_val[list_parameter_mo_val_key]).upper().strip().replace("'",
                                                                                                             '').replace(
                    "NULL", 'NONE')
                if item_count == 1:
                    first_Item = list_parameter_mo_val_key + '(' + value_site + ')'
                if cell_id > 0:
                    self.ws2["B" + str(index)] = str(cell_id) + '-' + first_Item
                else:
                    self.ws2["B" + str(index)] = first_Item

                if value_site == 'NONE_COLUMN_FOUND':
                    self.ws2["C" + str(index)] = mo_name
                    self.ws2["D" + str(index)] = list_parameter_mo_val_key
                    self.ws2["E" + str(index)] = list_parameter_mo_val_value
                    self.ws2["F" + str(index)] = value_site
                    self.ws2["G" + str(index)] = "Fail"
                    self.ws2["H" + str(index)] = "Fail Column Not found"
                else:
                    if list_parameter_mo_val_value == value_site:
                        self.ws2["C" + str(index)] = mo_name
                        self.ws2["D" + str(index)] = list_parameter_mo_val_key
                        self.ws2["E" + str(index)] = list_parameter_mo_val_value
                        self.ws2["F" + str(index)] = value_site
                        self.ws2["G" + str(index)] = "Correct"
                        self.ws2["H" + str(index)] = "Correct Value Configurate"
                    else:
                        self.ws2["C" + str(index)] = mo_name
                        self.ws2["D" + str(index)] = list_parameter_mo_val_key
                        self.ws2["E" + str(index)] = list_parameter_mo_val_value
                        self.ws2["F" + str(index)] = value_site
                        self.ws2["G" + str(index)] = "Fail"
                        self.ws2["H" + str(index)] = "Fail Value Configurate"

                index = index + 1
        return index

    def validate_utranncell(self,
                            p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                            list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        Cell_id_DF4G = ""
        row_identificated = ""
        list_parameter_df: list = ["Local Cell ID", "Rnc Id", "Rnc Cell Id", "Blind Handover Priority",
                                   "Local Cell Name",
                                   "Neighbour Cell Name", "Mobile Country Code", "Mobile Network Code"]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-UTRANNCELL"
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    RNC_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    RNC_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Blind_handover_priority_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Local_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Neighbour_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    list_value_correct = []

                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(RNC_ID_df)
                    list_value_correct.append(RNC_cell_ID_df)
                    list_value_correct.append(Blind_handover_priority_df)
                    list_value_correct.append(Local_cell_name_df)
                    list_value_correct.append(Neighbour_cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df and Local_cell_ID_df == Cell_id_DF4G:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-UTRANNCELL", list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):
                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Neighbour_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=15).value).strip().upper()
                                Local_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=14).value).strip().upper()

                                Local_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                            if eNodeB_Name_site == value_eNodeB_Name_df and Neighbour_cell_name_site == Neighbour_cell_name_df \
                                    and Local_cell_name_site == Local_cell_name_df and Local_cell_ID_site == Local_cell_ID_df:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                RNC_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                RNC_cell_ID_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Blind_handover_priority_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                # Local_cell_name_site = str(excel_sheet_site.cell(row=u, column=14).value).strip().upper()
                                # Neighbour_cell_name_site=str(excel_sheet_site.cell(row=u, column=15).value).strip().upper
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(RNC_ID_site)
                                list_value_current.append(RNC_cell_ID_site)
                                list_value_current.append(Blind_handover_priority_site)
                                list_value_current.append(Local_cell_name_site)
                                list_value_current.append(Neighbour_cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_utranexternalcell(self,
                                   p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                   list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        row_identificated = ""
        list_parameter_df: list = ["Rnc Id", "Rnc Cell Id", "Downlink Uarfcn", "Uplink Uarfcn Configure Indicator",
                                   "Utran Cell Type Indicator", "Routing area code configure indicator",
                                   "Routing Area Code", "Primary Scrambling Code", "Location Area Code", "Cell Name",
                                   "Mobile Country Code", "Mobile Network Code"]
        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = [],
        list_value_current = []
        lis_value_message = []
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)
        ifv = 0
        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            ifv = ifv + 1
            if ifv > 1:
                break
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            row_identificated = eNodeB_Name_DF4G
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':
                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    # Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    RNC_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    RNC_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Downlink_UARFCN_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Uplink_UARFCN_configure_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    UTRAN_cell_type_indicator_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Routing_area_code_configure_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Routing_area_code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Primary_scrambling_code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    Location_area_code_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    Cell_name_df = str(p_excel_sheet_df.cell(row=i, column=13).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=14).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=15).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=16).value).strip().upper()
                    list_value_correct = []
                    row_identificated = Cell_name_df + "-" + RNC_cell_ID_df

                    # Agregar los valores configurados actual de Config DF
                    list_value_correct.append(RNC_ID_df)
                    list_value_correct.append(RNC_cell_ID_df)
                    list_value_correct.append(Downlink_UARFCN_df)
                    list_value_correct.append(Uplink_UARFCN_configure_indicator_df)
                    list_value_correct.append(UTRAN_cell_type_indicator_df)
                    list_value_correct.append(Routing_area_code_configure_indicator_df)
                    list_value_correct.append(Routing_area_code_df)
                    list_value_correct.append(Primary_scrambling_code_df)
                    list_value_correct.append(Location_area_code_df)
                    list_value_correct.append(Cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if "MOVISTAR" == RAN_df and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-UTRANEXTERNALCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):
                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]

                                cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=14).value).upper().strip()

                            if eNodeB_Name_site == value_eNodeB_Name_df and cell_name_site == Cell_name_df:

                                RNC_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                RNC_cell_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                Downlink_UARFCN_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Uplink_UARFCN_configure_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                UTRAN_cell_type_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Routing_area_code_configure_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                Routing_area_code_site = str(
                                    excel_sheet_site.cell(row=u, column=11).value).strip().upper()
                                Primary_scrambling_code_site = str(
                                    excel_sheet_site.cell(row=u, column=12).value).strip().upper()
                                Location_area_code_site = str(
                                    excel_sheet_site.cell(row=u, column=13).value).strip().upper()
                                Cell_name_site = str(excel_sheet_site.cell(row=u, column=14).value).strip().upper()
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()

                                list_value_current = []
                                # Agregar los valores configurados actual de Config site
                                list_value_current.append(RNC_ID_site)
                                list_value_current.append(RNC_cell_ID_site)
                                list_value_current.append(Downlink_UARFCN_site)
                                list_value_current.append(Uplink_UARFCN_configure_indicator_site)
                                list_value_current.append(UTRAN_cell_type_indicator_site)
                                list_value_current.append(Routing_area_code_configure_indicator_site)
                                list_value_current.append(Routing_area_code_site)
                                list_value_current.append(Primary_scrambling_code_site)
                                list_value_current.append(Location_area_code_site)
                                list_value_current.append(Cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_utrannfreq(self,
                            p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                            list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "Downlink UARFCN", "Minimum required quality level",
                                   "Uplink UARFCN indicator",
                                   "Reselection priority configure indicator", "Cell reselection priority",
                                   "Frequency Priority for Connected Mode"]
        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = [],
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)
        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-UTRANNFREQ"
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':
                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    Downlink_UARFCN_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Minimum_required_quality_level_df = str(
                        p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Uplink_UARFCN_indicator_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Reselection_priority_configure_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Cell_reselection_priority_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Frequency_Priority_for_Connected_Mode_df = str(
                        p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    list_value_correct = []

                    # Agregar los valores configurados actual de Config DF
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(Downlink_UARFCN_df)
                    list_value_correct.append(Minimum_required_quality_level_df)
                    list_value_correct.append(Uplink_UARFCN_indicator_df)
                    list_value_correct.append(Reselection_priority_configure_indicator_df)
                    list_value_correct.append(Cell_reselection_priority_df)
                    list_value_correct.append(Frequency_Priority_for_Connected_Mode_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-UTRANNFREQ", list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):
                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).upper().strip()
                                Downlink_UARFCN_site = str(excel_sheet_site.cell(row=u, column=3).value).upper().strip()

                            if eNodeB_Name_site == value_eNodeB_Name_df and Local_cell_ID_site == Local_cell_ID_df and Downlink_UARFCN_site == Downlink_UARFCN_df:

                                RNC_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                list_value_current = []
                                # Agregar los valores configurados actual de Config site

                                Minimum_required_quality_level_site = str(
                                    excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Uplink_UARFCN_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                Reselection_priority_configure_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Cell_reselection_priority_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                Frequency_Priority_for_Connected_Mode_site = str(
                                    excel_sheet_site.cell(row=u, column=20).value).strip().upper()

                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(Downlink_UARFCN_site)
                                list_value_current.append(Minimum_required_quality_level_site)
                                list_value_current.append(Uplink_UARFCN_indicator_site)
                                list_value_current.append(Reselection_priority_configure_indicator_site)
                                list_value_current.append(Cell_reselection_priority_site)
                                list_value_current.append(Frequency_Priority_for_Connected_Mode_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
                    # else:
                    # index_result_2 = index_result_2 + 1
                    # write_ws2_result_mo_parameter("VECINDAD-UTRANNFREQ", list_parameter_df, index_result_2)
                    # index_result_2 = write_ws2_result_value_correct_current(list_parameter_mo=list_parameter_mo,
                    #                                                        list_parameter_df=list_parameter_df,
                    #                                                        list_value_correct=list_value_correct,
                    #                                                        list_value_current=list_value_current,
                    #                                                        index=index_result_2,
                    #                                                        flat_fail=True,
                    #                                                        message="The [NE Name] :{} or "
                    #                                                                "[eNodeB Name] :{} incorrect"
                    #                                                        .format(value_Ne_Name_df,
                    #                                                                value_eNodeB_Name_df))
        return index_result_2

    def validate_eutraninterfreqncell(self,
                                      p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                      list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "eNodeB ID", "NCell ID", "Blind handover Priority",
                                   "Local cell name",
                                   "Neighbour Cell Name", "Mobile Country Code", "Mobile Network Code"]
        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANINTERFREQNCELL"
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    eNodeB_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    NCell_ID = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Blind_handover_priority_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Local_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Neighbour_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(eNodeB_ID_df)
                    list_value_correct.append(NCell_ID)
                    list_value_correct.append(Blind_handover_priority_df)
                    list_value_correct.append(Local_cell_name_df)
                    list_value_correct.append(Neighbour_cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANINTERFREQNCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Neighbour_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=14).value).strip().upper()
                                Local_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=13).value).strip().upper()

                                Local_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                            if eNodeB_Name_site == value_eNodeB_Name_df and Neighbour_cell_name_site == Neighbour_cell_name_df \
                                    and Local_cell_name_site == Local_cell_name_df and Local_cell_ID_site == Local_cell_ID_df:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                eNodeB_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                NCell_ID = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Blind_handover_priority_site = str(
                                    excel_sheet_site.cell(row=u, column=11).value).strip().upper()
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(eNodeB_ID_site)
                                list_value_current.append(NCell_ID)
                                list_value_current.append(Blind_handover_priority_site)
                                list_value_current.append(Local_cell_name_site)
                                list_value_current.append(Neighbour_cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_eutraninternfreq(self,
                                  p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                  list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "Downlink EARFCN", "Uplink EARFCN indicator",
                                   "Inter Frequency cell resel priority indicator",
                                   "Inter Frequency cell resel priority",
                                   "EUTRAN reselection time", "Speed dependent resel parameter configuring indicator",
                                   "Measurement bandwidth",
                                   "P Max configuring indicator"]

        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANINTERNFREQ"
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    Downlink_EARFCN_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Uplink_EARFCN_indicator_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Inter_Frequency_cell_resel_priority_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Inter_Frequency_cell_resel_priority_df = str(
                        p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    EUTRAN_reselection_time_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Speed_dependent_resel_parameter_configuring_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Measurement_bandwidth_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    P_Max_configuring_indicator_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    Q_Qual_Min_configuring_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=13).value).strip().upper()
                    Clasificacion_df = str(p_excel_sheet_df.cell(row=i, column=14).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=15).value).strip().upper()

                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(Downlink_EARFCN_df)
                    list_value_correct.append(Uplink_EARFCN_indicator_df)
                    list_value_correct.append(Inter_Frequency_cell_resel_priority_indicator_df)
                    list_value_correct.append(Inter_Frequency_cell_resel_priority_df)
                    list_value_correct.append(EUTRAN_reselection_time_df)
                    list_value_correct.append(Speed_dependent_resel_parameter_configuring_indicator_df)
                    list_value_correct.append(Measurement_bandwidth_df)
                    list_value_correct.append(P_Max_configuring_indicator_df)
                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANINTERNFREQ",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Local_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()

                                Downlink_EARFCN_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                            if str(eNodeB_Name_site).upper() == str(
                                    value_eNodeB_Name_df).upper() and Local_cell_ID_df == Local_cell_ID_site \
                                    and Downlink_EARFCN_df == Downlink_EARFCN_site:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                Downlink_EARFCN_site = str(excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Uplink_EARFCN_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                Inter_Frequency_cell_resel_priority_indicator_site = \
                                    str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Inter_Frequency_cell_resel_priority_site = \
                                    str(excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                EUTRAN_reselection_time_site = str(
                                    excel_sheet_site.cell(row=u, column=8).value).strip().upper()
                                Speed_dependent_resel_parameter_configuring_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Measurement_bandwidth_site = str(
                                    excel_sheet_site.cell(row=u, column=12).value).strip().upper()
                                P_Max_configuring_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=17).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(Downlink_EARFCN_site)
                                list_value_current.append(Uplink_EARFCN_indicator_site)
                                list_value_current.append(Inter_Frequency_cell_resel_priority_indicator_site)
                                list_value_current.append(Inter_Frequency_cell_resel_priority_site)
                                list_value_current.append(EUTRAN_reselection_time_site)
                                list_value_current.append(Speed_dependent_resel_parameter_configuring_indicator_site)
                                list_value_current.append(Measurement_bandwidth_site)
                                list_value_current.append(P_Max_configuring_indicator_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_eutranintrafreqncell(self,
                                      p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                      list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "eNodeB ID", "Cell ID",
                                   "Local cell name",
                                   "Neighbour cell name",
                                   "Mobile Country Code", "Mobile Network Code"]

        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        key_row = ""
        value_row = ""
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANINTRAFREQNCELL"
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    eNodeB_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Local_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Neighbour_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()

                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(eNodeB_ID_df)
                    list_value_correct.append(Cell_ID_df)
                    list_value_correct.append(Local_cell_name_df)
                    list_value_correct.append(Neighbour_cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANINTRAFREQNCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Local_ell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=12).value).strip().upper()

                                Neighbour_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=13).value).strip().upper()
                            if eNodeB_Name_site == value_eNodeB_Name_df and Local_cell_name_df == Local_ell_name_site \
                                    and Neighbour_cell_name_df == Neighbour_cell_name_site:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                eNodeB_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                Cell_ID_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(eNodeB_ID_site)
                                list_value_current.append(Cell_ID_site)
                                list_value_current.append(Local_ell_name_site)
                                list_value_current.append(Neighbour_cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current,
                                    sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position,
                                    position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")

        return index_result_2

    def validate_eutranexternalcell(self,
                                    p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                    list_parameter_mo: dict):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Mobile country code", "Mobile network code",
                                   "eNodeB ID", "Cell ID", "Downlink EARFCN", "Uplink EARFCN indicator",
                                   "Physical cell ID", "Tracking area code", "Cell name"]

        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANEXTERNALCELL"
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE':

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Mobile_country_code_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    Mobile_network_code_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    eNodeB_ID_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Downlink_EARFCN_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Uplink_EARFCN_indicator_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Physical_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Tracking_area_code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    Cell_name_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=13).value).strip().upper()

                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Mobile_country_code_df)
                    list_value_correct.append(Mobile_network_code_df)
                    list_value_correct.append(eNodeB_ID_df)
                    list_value_correct.append(Cell_ID_df)
                    list_value_correct.append(Downlink_EARFCN_df)
                    list_value_correct.append(Uplink_EARFCN_indicator_df)
                    list_value_correct.append(Physical_cell_ID_df)
                    list_value_correct.append(Tracking_area_code_df)
                    list_value_correct.append(Cell_name_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Ne_Name_DF4G == value_Ne_Name_df and eNodeB_Name_DF4G == value_eNodeB_Name_df and Cell_id_DF4G == Cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANEXTERNALCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Local_ell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=11).value).strip().upper()

                            if eNodeB_Name_site == value_eNodeB_Name_df and Cell_name_df == Local_ell_name_site:

                                Mobile_country_code_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                Mobile_network_code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                eNodeB_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                Cell_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                Downlink_EARFCN_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Uplink_EARFCN_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                Physical_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Tracking_area_code_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                Cell_name_site = str(excel_sheet_site.cell(row=u, column=11).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Mobile_country_code_site)
                                list_value_current.append(Mobile_network_code_site)
                                list_value_current.append(eNodeB_ID_site)
                                list_value_current.append(Cell_ID_site)
                                list_value_current.append(Downlink_EARFCN_site)
                                list_value_current.append(Uplink_EARFCN_indicator_site)
                                list_value_current.append(Physical_cell_ID_site)
                                list_value_current.append(Tracking_area_code_site)
                                list_value_current.append(Cell_name_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current,
                                    sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position,
                                    position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_sheet_df(self, excel_sheet_lte: Worksheet, workbook_site1: Workbook,
                          workbook_site2: Workbook, list_name_df: list, index_result_1: int, index_result_2: int):
        list_parameter = {}
        list_output = {}

        row_id_result_2: int = 0  # Identificador Generado para cada item de la segunda Hoja de resultado

        VP_ParameterID = ""  # Obtiene el ParameterID que se encuentra en el MOB
        VP_ParameterName = ""  # Obtiene el ParameterName que se encuentra en el MOB
        VP_FeatureValue = ""  # Obtiene el FeatureValue que se encuentra en el MOB

        for name_mo_df in list_name_df:
            Excel_sheet_DF = self.excel_document_DF[name_mo_df]
            list_parameter = get_parameter_by_mo(excel_sheet_lte, str(name_mo_df).upper())
            # validar que exista la hoja [MO] en Confidata si no existe enviar el RnpData_BTS3900
            index_result_1 = index_result_1 + 1
            ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                index_result_1,
                'CONFIDATA',
                name_mo_df,
                self.excel_NameWorkbook_CONF_SITE1,
                workbook_site1,
                True, False)

            if ValidateExist_MO:
                Excel_sheet_CONF_SITE1 = workbook_site1[name_mo_df]
                # Validamos cada uno de las hojas:
                if name_mo_df == "UTRANNCELL":
                    index_result_2 = self.validate_utranncell(p_excel_sheet_df=Excel_sheet_DF,
                                                              excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                              index_result_2=index_result_2,
                                                              list_parameter_mo=list_parameter)
                if name_mo_df == "UTRANEXTERNALCELL":
                    index_result_2 = self.validate_utranexternalcell(p_excel_sheet_df=Excel_sheet_DF,
                                                                     excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                     index_result_2=index_result_2,
                                                                     list_parameter_mo=list_parameter)
                if name_mo_df == "UTRANNFREQ":
                    index_result_2 = self.validate_utrannfreq(p_excel_sheet_df=Excel_sheet_DF,
                                                              excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                              index_result_2=index_result_2,
                                                              list_parameter_mo=list_parameter)
                if name_mo_df == "EUTRANINTERFREQNCELL":
                    index_result_2 = self.validate_eutraninterfreqncell(p_excel_sheet_df=Excel_sheet_DF,
                                                                        excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                        index_result_2=index_result_2,
                                                                        list_parameter_mo=list_parameter)
                if name_mo_df == "EUTRANINTERNFREQ":
                    index_result_2 = self.validate_eutraninternfreq(p_excel_sheet_df=Excel_sheet_DF,
                                                                    excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                    index_result_2=index_result_2,
                                                                    list_parameter_mo=list_parameter)
                if name_mo_df == "EUTRANINTRAFREQNCELL":
                    index_result_2 = self.validate_eutranintrafreqncell(p_excel_sheet_df=Excel_sheet_DF,
                                                                        excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                        index_result_2=index_result_2,
                                                                        list_parameter_mo=list_parameter)
                if name_mo_df == "EUTRANEXTERNALCELL":
                    index_result_2 = self.validate_eutranexternalcell(p_excel_sheet_df=Excel_sheet_DF,
                                                                      excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                      index_result_2=index_result_2,
                                                                      list_parameter_mo=list_parameter)
            else:
                index_result_1 = index_result_1 + 1
                ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                    index_result_1,
                    'RNPDATA',
                    name_mo_df,
                    self.excel_NameWorkbook_CONF_SITE2,
                    workbook_site2,
                    True, False)
                if ValidateExist_MO:
                    Excel_sheet_CONF_SITE2 = workbook_site2[name_mo_df]

                    if name_mo_df == "UTRANNCELL":
                        index_result_2 = self.validate_utranncell(p_excel_sheet_df=Excel_sheet_DF,
                                                                  excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                  index_result_2=index_result_2,
                                                                  list_parameter_mo=list_parameter)
                    if name_mo_df == "UTRANEXTERNALCELL":
                        index_result_2 = self.validate_utranexternalcell(p_excel_sheet_df=Excel_sheet_DF,
                                                                         excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                         index_result_2=index_result_2,
                                                                         list_parameter_mo=list_parameter)
                    if name_mo_df == "UTRANNFREQ":
                        index_result_2 = self.validate_utrannfreq(p_excel_sheet_df=Excel_sheet_DF,
                                                                  excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                  index_result_2=index_result_2,
                                                                  list_parameter_mo=list_parameter)
                    if name_mo_df == "EUTRANINTERFREQNCELL":
                        index_result_2 = self.validate_eutraninterfreqncell(p_excel_sheet_df=Excel_sheet_DF,
                                                                            excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                            index_result_2=index_result_2,
                                                                            list_parameter_mo=list_parameter)
                    if name_mo_df == "EUTRANINTERNFREQ":
                        index_result_2 = self.validate_eutraninternfreq(p_excel_sheet_df=Excel_sheet_DF,
                                                                        excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                        index_result_2=index_result_2,
                                                                        list_parameter_mo=list_parameter)
                    if name_mo_df == "EUTRANINTRAFREQNCELL":
                        index_result_2 = self.validate_eutranintrafreqncell(p_excel_sheet_df=Excel_sheet_DF,
                                                                            excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                            index_result_2=index_result_2,
                                                                            list_parameter_mo=list_parameter)
                    if name_mo_df == "EUTRANEXTERNALCELL":
                        index_result_2 = self.validate_eutranexternalcell(p_excel_sheet_df=Excel_sheet_DF,
                                                                          excel_sheet_site=Excel_sheet_CONF_SITE2,
                                                                          index_result_2=index_result_2,
                                                                          list_parameter_mo=list_parameter)

        list_output["index_result_1"] = index_result_1
        list_output["index_result_2"] = index_result_2
        return list_output

    def write_ws2_sheet_others_value_correct_current(self, Cell_id_DF4G: int, Enode: str, list_parameter_mo: dict,
                                                     list_parameter_site: dict,
                                                     index: int,
                                                     mo_name: str):

        LTE_700Cell: bool = False
        parameter_contador = 0
        for key_site, key_site_value in list_parameter_site.items():
            list_value_site = dict(key_site_value)
            parameter_contador = 0
            LTE_700Cell = False
            LTE_700_or_LTE_AWS = 0
            # if len(list_parameter_mo) == 1:
            for key_parameter_mo, key_parameter_mo_value in list_parameter_mo.items():
                key_parameter_mo_value = str(key_parameter_mo_value).replace("'", '').replace("NULL", 'NONE').replace(
                    '"',
                    "")
                value_site = key_site_value[key_parameter_mo]
                parameter_contador = parameter_contador + 1
                parameter_name = str(key_parameter_mo).strip().title()
                parameter_name = str(parameter_name[:-len(str(parameter_contador))]).strip().title()
                if Cell_id_DF4G > 0:
                    self.ws2["B" + str(index)] = str(Cell_id_DF4G) + '-' + mo_name
                else:
                    self.ws2["B" + str(index)] = mo_name + '-' + Enode
                if key_parameter_mo_value.find('|') >= 0:
                    value_solo_un_valor = key_site_value[key_parameter_mo].replace("'", '').replace("NULL",
                                                                                                    'NONE').replace(
                        '"', "")
                    self.ws2["C" + str(index)] = mo_name
                    self.ws2["D" + str(index)] = parameter_name
                    self.ws2["E" + str(index)] = key_parameter_mo_value
                    self.ws2["F" + str(index)] = value_solo_un_valor
                    if LTE_700Cell is False:
                        if key_parameter_mo_value.strip() == "LTE_700 Cell | LTE_AWS Cell".strip():
                            LTE_700Cell = True
                            if convert_string_to_int(value_solo_un_valor):
                                if 100 >= int(value_solo_un_valor) >= 91:
                                    self.ws2["G" + str(index)] = "Correct"
                                    self.ws2["H" + str(index)] = "Correct Value Configurate LTE_700 Cell"
                                    LTE_700_or_LTE_AWS = 1
                                elif 103 >= int(value_solo_un_valor) >= 101:
                                    self.ws2["G" + str(index)] = "Correct"
                                    self.ws2["H" + str(index)] = "Correct Value Configurate LTE_AWS Cell"
                                    LTE_700_or_LTE_AWS = 2
                                else:
                                    self.ws2["G" + str(index)] = "Incorrect"
                                    self.ws2["H" + str(index)] = "Fail Value Configurate"
                                    LTE_700_or_LTE_AWS = -1
                            else:
                                self.ws2["G" + str(index)] = "Fail"
                                self.ws2["H" + str(index)] = "Fail Value Configurate"
                                LTE_700_or_LTE_AWS = -1
                    else:
                        Value_parameter_mo_value_split = str(key_parameter_mo_value).split('|')
                        Value_current_split = ""
                        if len(Value_parameter_mo_value_split) == 2:
                            if LTE_700_or_LTE_AWS == 1:
                                Value_current_split = Value_parameter_mo_value_split[0]
                            elif LTE_700_or_LTE_AWS == 2:
                                Value_current_split = Value_parameter_mo_value_split[1]
                            elif LTE_700_or_LTE_AWS == -1:
                                Value_current_split = "LTE_700_LTE_AWS_FOUND"
                            if Value_current_split == "LTE_700_LTE_AWS_FOUND":
                                self.ws2["G" + str(index)] = "Fail"
                                self.ws2["H" + str(index)] = "Fail Value Configurate LTE_700 or LTE_AWS Incorrect"
                            else:
                                texto_value_find = ""
                                item_count_split = 0
                                if Value_current_split.find(':') > 0:
                                    texto_value_find = texto_value_find + "|" + Value_current_split.strip()
                                    split_value = Value_current_split.strip().split(':')
                                    if len(split_value) == 2:
                                        self.ws2["C" + str(index)] = mo_name
                                        self.ws2["D" + str(index)] = parameter_name
                                        self.ws2["E" + str(index)] = Value_current_split

                                        value_parameter_confg = str(split_value[0]).strip()
                                        value_parameter_confg_value = str(split_value[1]).strip().upper()
                                        value_parameter_confg_value_convert = ""
                                        if value_parameter_confg_value == "ON":
                                            value_parameter_confg_value_convert = "1"
                                        elif value_parameter_confg_value == "OFF":
                                            value_parameter_confg_value_convert = "0"
                                        else:
                                            value_parameter_confg_value_convert = value_parameter_confg_value
                                            # hacer un split a los valos configurados actualmente,
                                        flat_Encontrado = False
                                        value_split_currect_data = str(value_site).split('&')
                                        for value_split_currect_val_data_value in value_split_currect_data:
                                            split_value_current_site = value_split_currect_val_data_value.split('-')
                                            if len(split_value_current_site) == 2:
                                                value_parameter_confg_site = split_value_current_site[0]
                                                value_parameter_confg_value_convert_site = split_value_current_site[1]
                                                if value_parameter_confg.upper().strip() == value_parameter_confg_site.upper().strip() and \
                                                        value_parameter_confg_value_convert.upper().strip() == value_parameter_confg_value_convert_site.upper().strip():
                                                    flat_Encontrado = True
                                                    item_count_split = item_count_split + 1
                                                    break

                                                else:
                                                    flat_Encontrado = False

                                        if flat_Encontrado is True:
                                            texto_value_find = texto_value_find + '(' + 'correct)'
                                        else:
                                            texto_value_find = texto_value_find + '(' + 'incorrect)'

                                        if 1 == item_count_split:
                                            self.ws2["F" + str(index)] = texto_value_find
                                            self.ws2["G" + str(index)] = "Correct"
                                            self.ws2["H" + str(index)] = "Correct Value Configurate"
                                        else:
                                            self.ws2["F" + str(index)] = texto_value_find
                                            self.ws2["G" + str(index)] = "Fail"
                                            self.ws2["H" + str(index)] = "Fail Value Configurate"

                                else:
                                    if str(Value_current_split).strip().upper() == value_solo_un_valor.strip().upper():
                                        self.ws2["G" + str(index)] = "Correct"
                                        self.ws2["H" + str(index)] = "Correct Value Configurate"
                                    else:
                                        self.ws2["G" + str(index)] = "Fail"
                                        self.ws2["H" + str(index)] = "Fail Value Configurate"

                elif key_parameter_mo_value.find(':') >= 0 and key_parameter_mo_value.find('|') == -1:
                    # Un split de los valores que se encuentran el LTE

                    texto_value_find = ""
                    Value_split_lte = str(key_parameter_mo_value).split(',')
                    len_item_split_lte = len(Value_split_lte)
                    item_count_split = 0
                    self.ws2["C" + str(index)] = mo_name
                    self.ws2["D" + str(index)] = parameter_name
                    self.ws2["E" + str(index)] = key_parameter_mo_value

                    for item_split_value in Value_split_lte:
                        texto_value_find = texto_value_find + "|" + item_split_value
                        split_value = item_split_value.split(':')
                        if len(split_value) == 2:
                            value_parameter_confg = str(split_value[0]).strip()
                            value_parameter_confg_value = str(split_value[1]).strip().upper()
                            value_parameter_confg_value_convert = ""
                            if value_parameter_confg_value == "ON":
                                value_parameter_confg_value_convert = "1"
                            elif value_parameter_confg_value == "OFF":
                                value_parameter_confg_value_convert = "0"
                            else:
                                value_parameter_confg_value_convert = value_parameter_confg_value

                            # hacer un split a los valos configurados actualmente,
                            flat_Encontrado = False
                            value_split_currect_data = str(value_site).split('&')
                            for value_split_currect_val_data_value in value_split_currect_data:
                                split_value_current_site = value_split_currect_val_data_value.split('-')
                                if len(split_value_current_site) == 2:
                                    value_parameter_confg_site = split_value_current_site[0]
                                    value_parameter_confg_value_convert_site = split_value_current_site[1]
                                    if value_parameter_confg.upper().strip() == value_parameter_confg_site.upper().strip() and \
                                            value_parameter_confg_value_convert.upper().strip() == value_parameter_confg_value_convert_site.upper().strip():
                                        flat_Encontrado = True
                                        item_count_split = item_count_split + 1
                                        break

                                    else:
                                        flat_Encontrado = False

                            if flat_Encontrado is True:
                                texto_value_find = texto_value_find + '(' + 'correct)'
                            else:
                                texto_value_find = texto_value_find + '(' + 'incorrect)'

                    if len_item_split_lte == item_count_split:
                        self.ws2["F" + str(index)] = texto_value_find
                        self.ws2["G" + str(index)] = "Correct"
                        self.ws2["H" + str(index)] = "Correct Value Configurate"
                    else:
                        self.ws2["F" + str(index)] = texto_value_find
                        self.ws2["G" + str(index)] = "Fail"
                        self.ws2["H" + str(index)] = "Fail Value Configurate"

                elif key_parameter_mo_value.find('&') >= 0:
                    print('')
                else:
                    value_solo_un_valor = key_site_value[key_parameter_mo].replace("'", '').replace("NULL",
                                                                                                    'NONE').replace(
                        '"', "")
                    if len(value_solo_un_valor) == 0:
                        value_solo_un_valor = 'NONE'
                    self.ws2["C" + str(index)] = mo_name
                    self.ws2["D" + str(index)] = parameter_name
                    self.ws2["E" + str(index)] = key_parameter_mo_value
                    self.ws2["F" + str(index)] = value_solo_un_valor

                    if str(key_parameter_mo_value).upper() == "L700_Cell".upper():
                        if convert_string_to_int(value_solo_un_valor):
                            if 100 >= int(value_solo_un_valor) >= 91:
                                self.ws2["G" + str(index)] = "Correct"
                                self.ws2["H" + str(index)] = "Correct Value Configurate"
                            else:
                                self.ws2["G" + str(index)] = "Fail"
                                self.ws2["H" + str(index)] = "Fail Value Configurate"
                        else:
                            self.ws2["G" + str(index)] = "Fail"
                            self.ws2["H" + str(index)] = "Fail Value Configurate"

                    else:
                        if str(value_solo_un_valor).upper().strip() == key_parameter_mo_value.upper().strip():
                            self.ws2["G" + str(index)] = "Correct"
                            self.ws2["H" + str(index)] = "Correct Value Configurate"
                        else:
                            self.ws2["G" + str(index)] = "Fail"
                            self.ws2["H" + str(index)] = "Fail Value Configurate"
                index = index + 1
        return index

    def validate_sheet_group_case_1(self, excel_sheet_lte: Worksheet, workbook_site1: Workbook,
                                    workbook_site2: Workbook, workbook_site3: Workbook, list_name_df: list,
                                    index_result_1: int, index_result_2: int):
        list_parameter = {}
        list_output = {}

        row_id_result_2: int = 0  # Identificador Generado para cada item de la segunda Hoja de resultado

        VP_ParameterID = ""  # Obtiene el ParameterID que se encuentra en el MOB
        VP_ParameterName = ""  # Obtiene el ParameterName que se encuentra en el MOB
        VP_FeatureValue = ""  # Obtiene el FeatureValue que se encuentra en el MOB
        position_id_cellId = 0  # validar si el grupo tiene CEllID
        for name_mo_df in list_name_df:
            # Excel_sheet_DF = excel_document_DF[name_mo_df]
            #
            index_result_1 = index_result_1 + 1
            ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                index_result_1,
                'CONFIDATA',
                name_mo_df,
                self.excel_NameWorkbook_CONF_SITE1,
                workbook_site1,
                True, False)
            if ValidateExist_MO:
                Excel_sheet_CONF_SITE1 = workbook_site1[name_mo_df]

                list_parameter = {}
                if str(name_mo_df).upper().strip() == "VOICEAMRCONTROL".upper().strip():
                    new_parameter_dist = {}
                    new_parameter_val = {}
                    list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                    Position_CellId = get_position_new_parameter(["Local Cell Id"], Excel_sheet_CONF_SITE1)
                    position_id_cellId = Position_CellId["Col_1"]
                    listNewParamater = group_case_1_get_parameter_by_id_group(list_parameter, position_id_cellId)

                    cont = 0
                    # replazamos las cabezeras
                    for list_parameter_key, list_parameter_value in listNewParamater.items():
                        list_parameter_dist_value = dict(list_parameter_value)
                        cont = cont + 1
                        new_parameter_val = {}
                        for key_new_parameter_val, value_new_parameter_val in list_parameter_dist_value.items():
                            if str(
                                    key_new_parameter_val).upper().strip() == "Voice AMR Control Parameter Group ID".upper().strip():
                                new_parameter_val["Voice Rate Control Parameter Group ID"] = value_new_parameter_val
                            if str(key_new_parameter_val).upper().strip() == "High AMR Coding Mode".upper().strip():
                                new_parameter_val["High Rate Coding Mode"] = value_new_parameter_val
                            if str(key_new_parameter_val).upper().strip() == "Low AMR Coding Mode".upper().strip():
                                new_parameter_val["Low Rate Coding Mode"] = value_new_parameter_val
                            if str(
                                    key_new_parameter_val).upper().strip() == "Packet Loss Rate Thd for Decreasing AMR".upper().strip():
                                new_parameter_val["Packet Loss Rate Thd for Decreasing"] = value_new_parameter_val
                            if str(
                                    key_new_parameter_val).upper().strip() == "Packet Loss Rate Thd for Increasing AMR".upper().strip():
                                new_parameter_val["Packet Loss Rate Threshold for Increasing"] = value_new_parameter_val
                            if str(
                                    key_new_parameter_val).upper().strip() == "RLC Segment Num Thd for Decreasing AMR".upper().strip():
                                new_parameter_val["RLC Segment Num Thd for Decreasing"] = value_new_parameter_val
                            if str(
                                    key_new_parameter_val).upper().strip() == "RLC Segment Num Thd for Increasing AMR".upper().strip():
                                new_parameter_val["RLC Segment Num Thd for Increasing"] = value_new_parameter_val
                        if position_id_cellId > 0:
                            new_parameter_val["Local Cell Id"] = 0
                        new_parameter_dist["key" + str(cont)] = new_parameter_val
                    listNewParamater = new_parameter_dist
                    listNewParamater_position = group_case_1_get_get_position_by_id_group(
                        list_parameter=listNewParamater,
                        sheet_name_site=Excel_sheet_CONF_SITE1)
                else:
                    list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                    Position_CellId = get_position_new_parameter(["Local Cell ID"], Excel_sheet_CONF_SITE1)
                    position_id_cellId = Position_CellId["Col_1"]
                    listNewParamater = group_case_1_get_parameter_by_id_group(list_parameter, position_id_cellId)
                    ## tengo que validar si existe el ["Local Cell ID"]

                    listNewParamater_position = group_case_1_get_get_position_by_id_group(
                        list_parameter=listNewParamater,

                        sheet_name_site=Excel_sheet_CONF_SITE1)
                # se debe duplicar el valor dependiendo de si existe
                if position_id_cellId == -1:
                    key = 0
                    col = 0
                    row_Evaluate = True
                    row_evaluate_val: int = 0
                    list_parameter_value_site_dic = {}
                    list_parameter_value_site = {}

                    for list_key, lis_value in listNewParamater.items():
                        key = key + 1
                        list_parameter_value_dic = dict(lis_value)
                        list_case_position = listNewParamater_position["key_" + str(key)]
                        CS1_TYPE1_SEARCH_ROW_VALUE = ""
                        col = 0
                        itemVal = 0
                        m = 0
                        RowId = 0
                        indexConf1 = 0
                        row_Evaluate = False
                        list_parameter_value_site = {}
                        # Obtener los valores de cada uno de los parametros que se encuentran en el Site
                        for list_key_parameter, list_key_parameter_value in list_parameter_value_dic.items():
                            col = col + 1
                            col_position = list_case_position["Col_" + str(col)]
                            if col_position > 0:
                                if row_Evaluate is False:
                                    for m in range(3, Excel_sheet_CONF_SITE1.max_row + 1):
                                        CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                            Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                        row=m).value).strip()
                                        if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                            if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == str(
                                                    list_key_parameter_value).upper().strip():
                                                list_parameter_value_site[
                                                    list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                row_evaluate_val = m
                                                itemVal = itemVal + 1
                                                row_Evaluate = True
                                                break
                                else:
                                    if row_evaluate_val > 0:
                                        CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                            Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                        row=row_evaluate_val).value).strip()
                                        list_parameter_value_site[list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                            else:
                                list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                        list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                    index_result_2 = self.write_ws2_group_case_1_value_correct_current(cell_id=0,
                                                                                       list_parameter_mo=listNewParamater,
                                                                                       list_parameter_site=list_parameter_value_site_dic,
                                                                                       index=index_result_2,
                                                                                       mo_name=name_mo_df)
                else:
                    item_row_df4g = get_df4g_row_site(
                        self.excel_document_DF["DF 4G"])  ## recorrer por cada Cell ID que se encuentre en el DF 4G
                    for item_parameter, value_row in item_row_df4g.items():
                        CellId_df4g = value_row["CellId_df4g"]
                        key = 0
                        col = 0
                        row_Evaluate = True
                        row_evaluate_val: int = 0
                        list_parameter_value_site_dic = {}
                        list_parameter_value_site = {}

                        for list_key, lis_value in listNewParamater.items():
                            key = key + 1
                            list_parameter_value_dic = dict(lis_value)
                            list_case_position = listNewParamater_position["key_" + str(key)]
                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                            col = 0
                            itemVal = 0
                            m = 0
                            RowId = 0
                            indexConf1 = 0
                            row_Evaluate = False
                            list_parameter_value_site = {}
                            # Obtener los valores de cada uno de los parametros que se encuentran en el Site
                            for list_key_parameter, list_key_parameter_value in list_parameter_value_dic.items():
                                col = col + 1
                                col_position = list_case_position["Col_" + str(col)]
                                if col_position > 0:
                                    if row_Evaluate is False:
                                        for m in range(3, Excel_sheet_CONF_SITE1.max_row + 1):
                                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                            row=m).value).strip()
                                            CellIdIndex = str(
                                                Excel_sheet_CONF_SITE1.cell(column=position_id_cellId,
                                                                            row=m).value).strip()
                                            if len(
                                                    CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                                if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == str(
                                                        list_key_parameter_value).upper().strip() \
                                                        and CellId_df4g == CellIdIndex:
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                    row_evaluate_val = m
                                                    itemVal = itemVal + 1
                                                    row_Evaluate = True
                                                    break
                                    else:
                                        if row_evaluate_val > 0:
                                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                            row=row_evaluate_val).value).strip()
                                            list_parameter_value_site[list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                else:
                                    list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                            list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                        index_result_2 = self.write_ws2_group_case_1_value_correct_current(cell_id=int(CellId_df4g),
                                                                                           list_parameter_mo=listNewParamater,
                                                                                           list_parameter_site=list_parameter_value_site_dic,
                                                                                           index=index_result_2,
                                                                                           mo_name=name_mo_df)
            else:
                index_result_1 = index_result_1 + 1
                ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                    index_result_1,
                    'RNPDATA',
                    name_mo_df,
                    self.excel_NameWorkbook_CONF_SITE2,
                    workbook_site2,
                    True, False)
                if ValidateExist_MO:
                    Excel_sheet_CONF_SITE2 = workbook_site2[name_mo_df]
                    list_parameter = {}
                    if str(name_mo_df).upper().strip() == "VOICEAMRCONTROL".upper().strip():
                        new_parameter_dist = {}
                        new_parameter_val = {}
                        list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                        Position_CellId = get_position_new_parameter(["Local Cell Id"], Excel_sheet_CONF_SITE2)
                        position_id_cellId = Position_CellId["Col_1"]
                        listNewParamater = group_case_1_get_parameter_by_id_group(list_parameter, position_id_cellId)

                        cont = 0
                        # replazamos las cabezeras
                        for list_parameter_key, list_parameter_value in listNewParamater.items():
                            list_parameter_dist_value = dict(list_parameter_value)
                            cont = cont + 1
                            new_parameter_val = {}
                            for key_new_parameter_val, value_new_parameter_val in list_parameter_dist_value.items():
                                if str(
                                        key_new_parameter_val).upper().strip() == "Voice AMR Control Parameter Group ID".upper().strip():
                                    new_parameter_val["Voice Rate Control Parameter Group ID"] = value_new_parameter_val
                                if str(key_new_parameter_val).upper().strip() == "High AMR Coding Mode".upper().strip():
                                    new_parameter_val["High Rate Coding Mode"] = value_new_parameter_val
                                if str(key_new_parameter_val).upper().strip() == "Low AMR Coding Mode".upper().strip():
                                    new_parameter_val["Low Rate Coding Mode"] = value_new_parameter_val
                                if str(
                                        key_new_parameter_val).upper().strip() == "Packet Loss Rate Thd for Decreasing AMR".upper().strip():
                                    new_parameter_val["Packet Loss Rate Thd for Decreasing"] = value_new_parameter_val
                                if str(
                                        key_new_parameter_val).upper().strip() == "Packet Loss Rate Thd for Increasing AMR".upper().strip():
                                    new_parameter_val[
                                        "Packet Loss Rate Threshold for Increasing"] = value_new_parameter_val
                                if str(
                                        key_new_parameter_val).upper().strip() == "RLC Segment Num Thd for Decreasing AMR".upper().strip():
                                    new_parameter_val["RLC Segment Num Thd for Decreasing"] = value_new_parameter_val
                                if str(
                                        key_new_parameter_val).upper().strip() == "RLC Segment Num Thd for Increasing AMR".upper().strip():
                                    new_parameter_val["RLC Segment Num Thd for Increasing"] = value_new_parameter_val
                            if position_id_cellId > 0:
                                new_parameter_val["Local Cell Id"] = 0
                            new_parameter_dist["key" + str(cont)] = new_parameter_val
                        listNewParamater = new_parameter_dist
                        listNewParamater_position = group_case_1_get_get_position_by_id_group(
                            list_parameter=listNewParamater,
                            sheet_name_site=Excel_sheet_CONF_SITE2)
                    else:
                        list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                        Position_CellId = get_position_new_parameter(["Local Cell ID"], Excel_sheet_CONF_SITE2)
                        position_id_cellId = Position_CellId["Col_1"]
                        listNewParamater = group_case_1_get_parameter_by_id_group(list_parameter, position_id_cellId)
                        ## tengo que validar si existe el ["Local Cell ID"]

                        listNewParamater_position = group_case_1_get_get_position_by_id_group(
                            list_parameter=listNewParamater,

                            sheet_name_site=Excel_sheet_CONF_SITE2)
                    # se debe duplicar el valor dependiendo de si existe
                    if position_id_cellId == -1:
                        key = 0
                        col = 0
                        row_Evaluate = True
                        row_evaluate_val: int = 0
                        list_parameter_value_site_dic = {}
                        list_parameter_value_site = {}

                        for list_key, lis_value in listNewParamater.items():
                            key = key + 1
                            list_parameter_value_dic = dict(lis_value)
                            list_case_position = listNewParamater_position["key_" + str(key)]
                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                            col = 0
                            itemVal = 0
                            m = 0
                            RowId = 0
                            indexConf1 = 0
                            row_Evaluate = False
                            list_parameter_value_site = {}
                            # Obtener los valores de cada uno de los parametros que se encuentran en el Site
                            for list_key_parameter, list_key_parameter_value in list_parameter_value_dic.items():
                                col = col + 1
                                col_position = list_case_position["Col_" + str(col)]
                                if col_position > 0:
                                    if row_Evaluate is False:
                                        for m in range(3, Excel_sheet_CONF_SITE2.max_row + 1):
                                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                            Excel_sheet_CONF_SITE2 = str(
                                                Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                            row=m).value).strip()
                                            if len(
                                                    CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                                if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == str(
                                                        list_key_parameter_value).upper().strip():
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                    row_evaluate_val = m
                                                    itemVal = itemVal + 1
                                                    row_Evaluate = True
                                                    break
                                    else:
                                        if row_evaluate_val > 0:
                                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                            row=row_evaluate_val).value).strip()
                                            list_parameter_value_site[list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                else:
                                    list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                            list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                        index_result_2 = self.write_ws2_group_case_1_value_correct_current(cell_id=0,
                                                                                           list_parameter_mo=listNewParamater,
                                                                                           list_parameter_site=list_parameter_value_site_dic,
                                                                                           index=index_result_2,
                                                                                           mo_name=name_mo_df)
                    else:
                        item_row_df4g = get_df4g_row_site(
                            self.excel_document_DF["DF 4G"])  ## recorrer por cada Cell ID que se encuentre en el DF 4G
                        for item_parameter, value_row in item_row_df4g.items():
                            CellId_df4g = value_row["CellId_df4g"]
                            key = 0
                            col = 0
                            row_Evaluate = True
                            row_evaluate_val: int = 0
                            list_parameter_value_site_dic = {}
                            list_parameter_value_site = {}

                            for list_key, lis_value in listNewParamater.items():
                                key = key + 1
                                list_parameter_value_dic = dict(lis_value)
                                list_case_position = listNewParamater_position["key_" + str(key)]
                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                col = 0
                                itemVal = 0
                                m = 0
                                RowId = 0
                                indexConf1 = 0
                                row_Evaluate = False
                                list_parameter_value_site = {}
                                # Obtener los valores de cada uno de los parametros que se encuentran en el Site
                                for list_key_parameter, list_key_parameter_value in list_parameter_value_dic.items():
                                    col = col + 1
                                    col_position = list_case_position["Col_" + str(col)]
                                    if col_position > 0:
                                        if row_Evaluate is False:
                                            for m in range(3, Excel_sheet_CONF_SITE2.max_row + 1):
                                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                                row=m).value).strip()
                                                CellIdIndex = str(
                                                    Excel_sheet_CONF_SITE2.cell(column=position_id_cellId,
                                                                                row=m).value).strip()
                                                if len(
                                                        CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                                    if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == str(
                                                            list_key_parameter_value).upper().strip() \
                                                            and CellId_df4g == CellIdIndex:
                                                        list_parameter_value_site[
                                                            list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                        row_evaluate_val = m
                                                        itemVal = itemVal + 1
                                                        row_Evaluate = True
                                                        break
                                        else:
                                            if row_evaluate_val > 0:
                                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                                row=row_evaluate_val).value).strip()
                                                list_parameter_value_site[
                                                    list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                    else:
                                        list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                                list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                            index_result_2 = self.write_ws2_group_case_1_value_correct_current(cell_id=int(CellId_df4g),
                                                                                               list_parameter_mo=listNewParamater,
                                                                                               list_parameter_site=list_parameter_value_site_dic,
                                                                                               index=index_result_2,
                                                                                               mo_name=name_mo_df)
                else:
                    index_result_1 = index_result_1 + 1
                    ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                        index_result_1,
                        'ConfigurationData_eNodeB',
                        name_mo_df,
                        self.excel_NameWorkbook_CONF_SITE3,
                        workbook_site3,
                        True, False)
                    if ValidateExist_MO:

                        Excel_sheet_CONF_SITE3 = workbook_site3[name_mo_df]
                        list_parameter = {}
                        if str(name_mo_df).upper().strip() == "VOICEAMRCONTROL".upper().strip():
                            new_parameter_dist = {}
                            new_parameter_val = {}
                            list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                            Position_CellId = get_position_new_parameter(["Local Cell Id"], Excel_sheet_CONF_SITE3)
                            position_id_cellId = Position_CellId["Col_1"]
                            listNewParamater = group_case_1_get_parameter_by_id_group(list_parameter,
                                                                                      position_id_cellId)

                            cont = 0
                            # replazamos las cabezeras
                            for list_parameter_key, list_parameter_value in listNewParamater.items():
                                list_parameter_dist_value = dict(list_parameter_value)
                                cont = cont + 1
                                new_parameter_val = {}
                                for key_new_parameter_val, value_new_parameter_val in list_parameter_dist_value.items():
                                    if str(
                                            key_new_parameter_val).upper().strip() == "Voice AMR Control Parameter Group ID".upper().strip():
                                        new_parameter_val[
                                            "Voice Rate Control Parameter Group ID"] = value_new_parameter_val
                                    if str(
                                            key_new_parameter_val).upper().strip() == "High AMR Coding Mode".upper().strip():
                                        new_parameter_val["High Rate Coding Mode"] = value_new_parameter_val
                                    if str(
                                            key_new_parameter_val).upper().strip() == "Low AMR Coding Mode".upper().strip():
                                        new_parameter_val["Low Rate Coding Mode"] = value_new_parameter_val
                                    if str(
                                            key_new_parameter_val).upper().strip() == "Packet Loss Rate Thd for Decreasing AMR".upper().strip():
                                        new_parameter_val[
                                            "Packet Loss Rate Thd for Decreasing"] = value_new_parameter_val
                                    if str(
                                            key_new_parameter_val).upper().strip() == "Packet Loss Rate Thd for Increasing AMR".upper().strip():
                                        new_parameter_val[
                                            "Packet Loss Rate Threshold for Increasing"] = value_new_parameter_val
                                    if str(
                                            key_new_parameter_val).upper().strip() == "RLC Segment Num Thd for Decreasing AMR".upper().strip():
                                        new_parameter_val[
                                            "RLC Segment Num Thd for Decreasing"] = value_new_parameter_val
                                    if str(
                                            key_new_parameter_val).upper().strip() == "RLC Segment Num Thd for Increasing AMR".upper().strip():
                                        new_parameter_val[
                                            "RLC Segment Num Thd for Increasing"] = value_new_parameter_val
                                if position_id_cellId > 0:
                                    new_parameter_val["Local Cell Id"] = 0
                                new_parameter_dist["key" + str(cont)] = new_parameter_val
                            listNewParamater = new_parameter_dist
                            listNewParamater_position = group_case_1_get_get_position_by_id_group(
                                list_parameter=listNewParamater,
                                sheet_name_site=Excel_sheet_CONF_SITE3)
                        else:
                            list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                            Position_CellId = get_position_new_parameter(["Local Cell ID"], Excel_sheet_CONF_SITE3)
                            position_id_cellId = Position_CellId["Col_1"]
                            listNewParamater = group_case_1_get_parameter_by_id_group(list_parameter,
                                                                                      position_id_cellId)
                            ## tengo que validar si existe el ["Local Cell ID"]

                            listNewParamater_position = group_case_1_get_get_position_by_id_group(
                                list_parameter=listNewParamater,

                                sheet_name_site=Excel_sheet_CONF_SITE3)
                        # se debe duplicar el valor dependiendo de si existe
                        if position_id_cellId == -1:
                            key = 0
                            col = 0
                            row_Evaluate = True
                            row_evaluate_val: int = 0
                            list_parameter_value_site_dic = {}
                            list_parameter_value_site = {}

                            for list_key, lis_value in listNewParamater.items():
                                key = key + 1
                                list_parameter_value_dic = dict(lis_value)
                                list_case_position = listNewParamater_position["key_" + str(key)]
                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                col = 0
                                itemVal = 0
                                m = 0
                                RowId = 0
                                indexConf1 = 0
                                row_Evaluate = False
                                list_parameter_value_site = {}
                                # Obtener los valores de cada uno de los parametros que se encuentran en el Site
                                for list_key_parameter, list_key_parameter_value in list_parameter_value_dic.items():
                                    col = col + 1
                                    col_position = list_case_position["Col_" + str(col)]
                                    if col_position > 0:
                                        if row_Evaluate is False:
                                            for m in range(3, Excel_sheet_CONF_SITE3.max_row + 1):
                                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                row=m).value).strip()
                                                if len(
                                                        CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                                    if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == str(
                                                            list_key_parameter_value).upper().strip():
                                                        list_parameter_value_site[
                                                            list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                        row_evaluate_val = m
                                                        itemVal = itemVal + 1
                                                        row_Evaluate = True
                                                        break
                                        else:
                                            if row_evaluate_val > 0:
                                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                row=row_evaluate_val).value).strip()
                                                list_parameter_value_site[
                                                    list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                    else:
                                        list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                                list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                            index_result_2 = self.write_ws2_group_case_1_value_correct_current(cell_id=0,
                                                                                               list_parameter_mo=listNewParamater,
                                                                                               list_parameter_site=list_parameter_value_site_dic,
                                                                                               index=index_result_2,
                                                                                               mo_name=name_mo_df)
                        else:
                            item_row_df4g = get_df4g_row_site(
                                self.excel_document_DF[
                                    "DF 4G"])  ## recorrer por cada Cell ID que se encuentre en el DF 4G
                            for item_parameter, value_row in item_row_df4g.items():
                                CellId_df4g = value_row["CellId_df4g"]
                                key = 0
                                col = 0
                                row_Evaluate = True
                                row_evaluate_val: int = 0
                                list_parameter_value_site_dic = {}
                                list_parameter_value_site = {}

                                for list_key, lis_value in listNewParamater.items():
                                    key = key + 1
                                    list_parameter_value_dic = dict(lis_value)
                                    list_case_position = listNewParamater_position["key_" + str(key)]
                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                    col = 0
                                    itemVal = 0
                                    m = 0
                                    RowId = 0
                                    indexConf1 = 0
                                    row_Evaluate = False
                                    list_parameter_value_site = {}
                                    # Obtener los valores de cada uno de los parametros que se encuentran en el Site
                                    for list_key_parameter, list_key_parameter_value in list_parameter_value_dic.items():
                                        col = col + 1
                                        col_position = list_case_position["Col_" + str(col)]
                                        if col_position > 0:
                                            if row_Evaluate is False:
                                                for m in range(3, Excel_sheet_CONF_SITE3.max_row + 1):
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                        Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                    row=m).value).strip()
                                                    CellIdIndex = str(
                                                        Excel_sheet_CONF_SITE3.cell(column=position_id_cellId,
                                                                                    row=m).value).strip()
                                                    if len(
                                                            CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                                        if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == str(
                                                                list_key_parameter_value).upper().strip() \
                                                                and CellId_df4g == CellIdIndex:
                                                            list_parameter_value_site[
                                                                list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                            row_evaluate_val = m
                                                            itemVal = itemVal + 1
                                                            row_Evaluate = True
                                                            break
                                            else:
                                                if row_evaluate_val > 0:
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                        Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                    row=row_evaluate_val).value).strip()
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                        else:
                                            list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                                    list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                                index_result_2 = self.write_ws2_group_case_1_value_correct_current(
                                    cell_id=int(CellId_df4g),
                                    list_parameter_mo=listNewParamater,
                                    list_parameter_site=list_parameter_value_site_dic,
                                    index=index_result_2,
                                    mo_name=name_mo_df)

        list_output["index_result_1"] = index_result_1
        list_output["index_result_2"] = index_result_2
        return list_output

    def validate_sheet_others(self, excel_sheet_lte: Worksheet, workbook_site1: Workbook,
                              workbook_site2: Workbook, workbook_site3: Workbook, list_name_df: list,
                              index_result_1: int,
                              index_result_2: int
                              ):
        list_parameter = {}
        list_output = {}

        row_id_result_2: int = 0  # Identificador Generado para cada item de la segunda Hoja de resultado

        VP_ParameterID = ""  # Obtiene el ParameterID que se encuentra en el MOB
        VP_ParameterName = ""  # Obtiene el ParameterName que se encuentra en el MOB
        VP_FeatureValue = ""  # Obtiene el FeatureValue que se encuentra en el MOB
        row_identificated = ""
        for name_mo_df in list_name_df:
            index_result_1 = index_result_1 + 1

            ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                index_result_1,
                'CONFIDATA',
                name_mo_df,
                self.excel_NameWorkbook_CONF_SITE1,
                workbook_site1,
                True, False)

            if ValidateExist_MO:
                Excel_sheet_CONF_SITE1 = workbook_site1[name_mo_df]
                list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                Position_CellId = get_position_new_parameter(["Local Cell Id"], Excel_sheet_CONF_SITE1)
                position_id_cellId = Position_CellId["Col_1"]
                list_parameter_index = str((len(list_parameter) + 1))
                if position_id_cellId > 0:
                    list_parameter["Local Cell Id" + list_parameter_index] = "0"
                list_Paramater_position = get_position_other(list_parameter=list_parameter,
                                                             sheet_name_site=Excel_sheet_CONF_SITE1)
                key = 0
                list_parameter_value_site_dic = {}
                list_parameter_value_site = {}
                list_key_parameter_1Value = {}
                position_parameter = dict(list_Paramater_position["key_1"])
                list_evaluate1Value = ["CSFALLBACKBLINDHOCFG", "NCELLPARACFG",
                                       "X2", "IRATNCELLCLASSMGT", "SCTPHOST",
                                       "USERPLANEHOST"]  # Evaluamos solo en valor conincidente.
                list_key_parameter_1Value["CSFALLBACKBLINDHOCFG"] = "Cn Operator Id|2|1"
                list_key_parameter_1Value["NCELLPARACFG"] = "Rat Type|3|2"
                list_key_parameter_1Value["X2"] = "X2 Id|2|1"
                list_key_parameter_1Value["IRATNCELLCLASSMGT"] = "Rat Type|2|1"
                list_key_parameter_1Value[
                    "SCTPHOST"] = "Sctp Host Id|2|1"  # NombreParametro|Posiciondel Site|Posicion del caberzera.
                list_key_parameter_1Value[
                    "USERPLANEHOST"] = "User Plane Host Id|2|1"  # NombreParametro|Posiciondel Site|Posicion del caberzera.
                item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
                Node_DF4G = ""
                # validar que exista [NE NAME or ENODENAME]
                Position_CellId_NE_ENODE_NAME1 = get_position_new_parameter(["Name"], Excel_sheet_CONF_SITE1)
                position_id_cellId_NE_ENODE_NAME1 = Position_CellId_NE_ENODE_NAME1["Col_1"]

                Position_CellId_NE_ENODE_NAME2 = get_position_new_parameter(["eNodeB Name"], Excel_sheet_CONF_SITE1)
                position_id_cellId_NE_ENODE_NAME2 = Position_CellId_NE_ENODE_NAME2["Col_1"]

                if position_id_cellId_NE_ENODE_NAME1 == 1 and position_id_cellId_NE_ENODE_NAME2 == -1:
                    Node_DF4G = item_row_df4g["site_1"]["NE_Name_df4g"]
                else:
                    Node_DF4G = item_row_df4g["site_1"]["eNodeB_Name_df4g"]

                if position_id_cellId == -1:
                    if name_mo_df in list_evaluate1Value:
                        key_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[0]
                        position_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[1]
                        position_row = str(list_key_parameter_1Value[name_mo_df]).split('|')[2]
                        val = list_parameter[key_value + position_row]
                        for m in range(3, Excel_sheet_CONF_SITE1.max_row + 1):
                            CS1_TYPE1_node = str(
                                Excel_sheet_CONF_SITE1.cell(column=1,
                                                            row=m).value).strip()
                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                Excel_sheet_CONF_SITE1.cell(column=int(position_value),
                                                            row=m).value).strip()
                            if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                if CS1_TYPE1_node.upper().strip() == Node_DF4G.upper().strip():
                                    if val == CS1_TYPE1_SEARCH_ROW_VALUE:
                                        key = key + 1
                                        col = 0
                                        list_parameter_value_site = {}
                                        for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                            col = col + 1
                                            col_position = position_parameter["Col_" + str(col)]
                                            if col_position > 0:
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                                row=m).value).strip()
                                                list_parameter_value_site[
                                                    list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                            else:
                                                list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                                        list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site

                    else:
                        for m in range(3, Excel_sheet_CONF_SITE1.max_row + 1):
                            CS1_TYPE1_SEARCH_ROW_VALUE = ""
                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                Excel_sheet_CONF_SITE1.cell(column=1,
                                                            row=m).value).strip()
                            if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                # and CS1_TYPE1_SEARCH_ROW_VALUE =="":
                                if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == Node_DF4G.upper().strip():
                                    key = key + 1
                                    col = 0

                                    list_parameter_value_site = {}
                                    for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                        col = col + 1
                                        col_position = position_parameter["Col_" + str(col)]
                                        if col_position > 0:
                                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                Excel_sheet_CONF_SITE1.cell(column=col_position, row=m).value).strip()
                                            list_parameter_value_site[list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                            # CellAlgoSwitch
                                        else:
                                            list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'

                            list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                    index_result_2 = self.write_ws2_sheet_others_value_correct_current(Cell_id_DF4G=0, Enode=Node_DF4G,
                                                                                       list_parameter_mo=list_parameter,
                                                                                       list_parameter_site=list_parameter_value_site_dic,
                                                                                       index=index_result_2,
                                                                                       mo_name=name_mo_df)
                else:
                    Mo_parameter_valueDF4G = {}
                    Cell_id_DF4G = ""
                    for item_row_parameter, item_row_parameter_value in item_row_df4g.items():
                        Cell_id_DF4G = item_row_parameter_value["CellId_df4g"]
                        list_parameter["Local Cell Id" + list_parameter_index] = Cell_id_DF4G
                        list_parameter_value_site_dic = {}
                        Mo_parameter_valueDF4G["CellId_" + str(Cell_id_DF4G)] = list_parameter
                        if name_mo_df in list_evaluate1Value:
                            key_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[0]
                            position_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[1]
                            position_row = str(list_key_parameter_1Value[name_mo_df]).split('|')[2]
                            val = list_parameter[key_value + position_row]
                            for m in range(3, Excel_sheet_CONF_SITE1.max_row + 1):
                                # position_id_cellId
                                Index_cell_id = str(
                                    Excel_sheet_CONF_SITE1.cell(column=int(position_id_cellId),
                                                                row=m).value).strip()
                                CS1_TYPE1_node = str(
                                    Excel_sheet_CONF_SITE1.cell(column=1,
                                                                row=m).value).strip()
                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                    Excel_sheet_CONF_SITE1.cell(column=int(position_value),
                                                                row=m).value).strip()
                                if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                    if CS1_TYPE1_node.upper().strip() == Node_DF4G.upper().strip() and Cell_id_DF4G == Index_cell_id:
                                        if val == CS1_TYPE1_SEARCH_ROW_VALUE:
                                            key = key + 1
                                            col = 0
                                            list_parameter_value_site = {}
                                            for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                col = col + 1
                                                col_position = position_parameter["Col_" + str(col)]
                                                if col_position > 0:
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                        Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                                    row=m).value).strip()
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                else:
                                                    list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                                            list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site

                        else:
                            list_parameter_value_site_dic = {}
                            key = 0
                            for m in range(3, Excel_sheet_CONF_SITE1.max_row + 1):
                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                    Excel_sheet_CONF_SITE1.cell(column=1,
                                                                row=m).value).strip()
                                Index_cell_id = str(
                                    Excel_sheet_CONF_SITE1.cell(column=int(position_id_cellId),
                                                                row=m).value).strip()
                                if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                    # and CS1_TYPE1_SEARCH_ROW_VALUE =="":
                                    if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == Node_DF4G.upper().strip() and Cell_id_DF4G == Index_cell_id:
                                        key = key + 1
                                        col = 0

                                        list_parameter_value_site = {}
                                        for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                            col = col + 1
                                            col_position = position_parameter["Col_" + str(col)]
                                            if col_position > 0:
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE1.cell(column=col_position,
                                                                                row=m).value).strip()
                                                list_parameter_value_site[
                                                    list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                # CellAlgoSwitch
                                            else:
                                                list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'

                                        list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                        index_result_2 = self.write_ws2_sheet_others_value_correct_current(
                            Cell_id_DF4G=int(Cell_id_DF4G),
                            Enode=Node_DF4G,
                            list_parameter_mo=list_parameter,
                            list_parameter_site=list_parameter_value_site_dic,
                            index=index_result_2,
                            mo_name=name_mo_df)

            else:
                index_result_1 = index_result_1 + 1
                ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                    index_result_1,
                    'RNPDATA',
                    name_mo_df,
                    self.excel_NameWorkbook_CONF_SITE2,
                    workbook_site2,
                    True, False)
                if ValidateExist_MO:
                    Excel_sheet_CONF_SITE2 = workbook_site1[name_mo_df]
                    list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                    Position_CellId = get_position_new_parameter(["Local Cell Id"], Excel_sheet_CONF_SITE2)
                    position_id_cellId = Position_CellId["Col_1"]
                    list_parameter_index = str((len(list_parameter) + 1))
                    if position_id_cellId > 0:
                        list_parameter["Local Cell Id" + list_parameter_index] = "0"
                    list_Paramater_position = get_position_other(list_parameter=list_parameter,
                                                                 sheet_name_site=Excel_sheet_CONF_SITE2)
                    key = 0
                    list_parameter_value_site_dic = {}
                    list_parameter_value_site = {}
                    list_key_parameter_1Value = {}
                    position_parameter = dict(list_Paramater_position["key_1"])
                    list_evaluate1Value = ["CSFALLBACKBLINDHOCFG", "NCELLPARACFG",
                                           "X2", "IRATNCELLCLASSMGT", "SCTPHOST",
                                           "USERPLANEHOST"]  # Evaluamos solo en valor conincidente.
                    list_key_parameter_1Value["CSFALLBACKBLINDHOCFG"] = "Cn Operator Id|2|1"
                    list_key_parameter_1Value["NCELLPARACFG"] = "Rat Type|3|2"
                    list_key_parameter_1Value["X2"] = "X2 Id|2|1"
                    list_key_parameter_1Value["IRATNCELLCLASSMGT"] = "Rat Type|2|1"
                    list_key_parameter_1Value[
                        "SCTPHOST"] = "Sctp Host Id|2|1"  # NombreParametro|Posiciondel Site|Posicion del caberzera.
                    list_key_parameter_1Value[
                        "USERPLANEHOST"] = "User Plane Host Id|2|1"  # NombreParametro|Posiciondel Site|Posicion del caberzera.
                    item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
                    Node_DF4G = ""
                    # validar que exista [NE NAME or ENODENAME]
                    Position_CellId_NE_ENODE_NAME1 = get_position_new_parameter(["Name"], Excel_sheet_CONF_SITE2)
                    position_id_cellId_NE_ENODE_NAME1 = Position_CellId_NE_ENODE_NAME1["Col_1"]

                    Position_CellId_NE_ENODE_NAME2 = get_position_new_parameter(["eNodeB Name"], Excel_sheet_CONF_SITE2)
                    position_id_cellId_NE_ENODE_NAME2 = Position_CellId_NE_ENODE_NAME2["Col_1"]

                    if position_id_cellId_NE_ENODE_NAME1 == 1 and position_id_cellId_NE_ENODE_NAME2 == -1:
                        Node_DF4G = item_row_df4g["site_1"]["NE_Name_df4g"]
                    else:
                        Node_DF4G = item_row_df4g["site_1"]["eNodeB_Name_df4g"]

                    if position_id_cellId == -1:
                        if name_mo_df in list_evaluate1Value:
                            key_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[0]
                            position_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[1]
                            position_row = str(list_key_parameter_1Value[name_mo_df]).split('|')[2]
                            val = list_parameter[key_value + position_row]
                            for m in range(3, Excel_sheet_CONF_SITE2.max_row + 1):
                                CS1_TYPE1_node = str(
                                    Excel_sheet_CONF_SITE2.cell(column=1,
                                                                row=m).value).strip()
                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                    Excel_sheet_CONF_SITE2.cell(column=int(position_value),
                                                                row=m).value).strip()
                                if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                    if CS1_TYPE1_node.upper().strip() == Node_DF4G.upper().strip():
                                        if val == CS1_TYPE1_SEARCH_ROW_VALUE:
                                            key = key + 1
                                            col = 0
                                            list_parameter_value_site = {}
                                            for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                col = col + 1
                                                col_position = position_parameter["Col_" + str(col)]
                                                if col_position > 0:
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                        Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                                    row=m).value).strip()
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                else:
                                                    list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'
                                            list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site

                        else:
                            for m in range(3, Excel_sheet_CONF_SITE2.max_row + 1):
                                CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                    Excel_sheet_CONF_SITE2.cell(column=1,
                                                                row=m).value).strip()
                                if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                    # and CS1_TYPE1_SEARCH_ROW_VALUE =="":
                                    if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == Node_DF4G.upper().strip():
                                        key = key + 1
                                        col = 0

                                        list_parameter_value_site = {}
                                        for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                            col = col + 1
                                            col_position = position_parameter["Col_" + str(col)]
                                            if col_position > 0:
                                                CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                    Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                                row=m).value).strip()
                                                list_parameter_value_site[
                                                    list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                # CellAlgoSwitch
                                            else:
                                                list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'

                                list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                        index_result_2 = self.write_ws2_sheet_others_value_correct_current(Cell_id_DF4G=0,
                                                                                           Enode=Node_DF4G,
                                                                                           list_parameter_mo=list_parameter,
                                                                                           list_parameter_site=list_parameter_value_site_dic,
                                                                                           index=index_result_2,
                                                                                           mo_name=name_mo_df)
                    else:
                        Cell_id_DF4G = ""
                        for item_row_parameter, item_row_parameter_value in item_row_df4g.items():
                            Cell_id_DF4G = item_row_parameter_value["CellId_df4g"]
                            list_parameter["Local Cell Id" + list_parameter_index] = Cell_id_DF4G
                            if name_mo_df in list_evaluate1Value:
                                key_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[0]
                                position_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[1]
                                position_row = str(list_key_parameter_1Value[name_mo_df]).split('|')[2]
                                val = list_parameter[key_value + position_row]
                                for m in range(3, Excel_sheet_CONF_SITE2.max_row + 1):
                                    # position_id_cellId
                                    Index_cell_id = str(
                                        Excel_sheet_CONF_SITE2.cell(column=int(position_id_cellId),
                                                                    row=m).value).strip()
                                    CS1_TYPE1_node = str(
                                        Excel_sheet_CONF_SITE2.cell(column=1,
                                                                    row=m).value).strip()
                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                        Excel_sheet_CONF_SITE2.cell(column=int(position_value),
                                                                    row=m).value).strip()
                                    if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                        if CS1_TYPE1_node.upper().strip() == Node_DF4G.upper().strip() and Cell_id_DF4G == Index_cell_id:
                                            if val == CS1_TYPE1_SEARCH_ROW_VALUE:
                                                key = key + 1
                                                col = 0
                                                list_parameter_value_site = {}
                                                for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                    col = col + 1
                                                    col_position = position_parameter["Col_" + str(col)]
                                                    if col_position > 0:
                                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                            Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                                        row=m).value).strip()
                                                        list_parameter_value_site[
                                                            list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                    else:
                                                        list_parameter_value_site[
                                                            list_key_parameter] = 'NONE_COLUMN_FOUND'
                                                list_parameter_value_site_dic[
                                                    "key_" + str(key)] = list_parameter_value_site

                            else:
                                list_parameter_value_site_dic = {}
                                key = 0
                                for m in range(3, Excel_sheet_CONF_SITE2.max_row + 1):
                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                        Excel_sheet_CONF_SITE2.cell(column=1,
                                                                    row=m).value).strip()
                                    Index_cell_id = str(
                                        Excel_sheet_CONF_SITE2.cell(column=int(position_id_cellId),
                                                                    row=m).value).strip()
                                    if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                        # and CS1_TYPE1_SEARCH_ROW_VALUE =="":
                                        if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == Node_DF4G.upper().strip() and Cell_id_DF4G == Index_cell_id:
                                            key = key + 1
                                            col = 0

                                            list_parameter_value_site = {}
                                            for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                col = col + 1
                                                col_position = position_parameter["Col_" + str(col)]
                                                if col_position > 0:
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                        Excel_sheet_CONF_SITE2.cell(column=col_position,
                                                                                    row=m).value).strip()
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                    # CellAlgoSwitch
                                                else:
                                                    list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'

                                    list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                            index_result_2 = self.write_ws2_sheet_others_value_correct_current(
                                Cell_id_DF4G=int(Cell_id_DF4G),
                                Enode=Node_DF4G,
                                list_parameter_mo=list_parameter,
                                list_parameter_site=list_parameter_value_site_dic,
                                index=index_result_2,
                                mo_name=name_mo_df)

                else:
                    index_result_1 = index_result_1 + 1
                    ValidateExist_MO = self.validate_exists_sheet_excel_by_workbook(
                        index_result_1,
                        'ConfigurationData_eNodeB',
                        name_mo_df,
                        self.excel_NameWorkbook_CONF_SITE3,
                        workbook_site3,
                        True, False)
                    if ValidateExist_MO:
                        Excel_sheet_CONF_SITE3 = workbook_site3[name_mo_df]
                        list_parameter = self.get_parameter_mo_by_group(excel_sheet_lte, str(name_mo_df).upper())
                        Position_CellId = get_position_new_parameter(["Local Cell Id"], Excel_sheet_CONF_SITE3)
                        position_id_cellId = Position_CellId["Col_1"]
                        list_parameter_index = str((len(list_parameter) + 1))
                        if position_id_cellId > 0:
                            list_parameter["Local Cell Id" + list_parameter_index] = "0"
                        list_Paramater_position = get_position_other(list_parameter=list_parameter,
                                                                     sheet_name_site=Excel_sheet_CONF_SITE3)
                        key = 0
                        list_parameter_value_site_dic = {}
                        list_parameter_value_site = {}
                        list_key_parameter_1Value = {}
                        position_parameter = dict(list_Paramater_position["key_1"])
                        list_evaluate1Value = ["CSFALLBACKBLINDHOCFG", "NCELLPARACFG",
                                               "X2", "IRATNCELLCLASSMGT", "SCTPHOST",
                                               "USERPLANEHOST"]  # Evaluamos solo en valor conincidente.
                        list_key_parameter_1Value["CSFALLBACKBLINDHOCFG"] = "Cn Operator Id|2|1"
                        list_key_parameter_1Value["NCELLPARACFG"] = "Rat Type|3|2"
                        list_key_parameter_1Value["X2"] = "X2 Id|2|1"
                        list_key_parameter_1Value["IRATNCELLCLASSMGT"] = "Rat Type|2|1"
                        list_key_parameter_1Value[
                            "SCTPHOST"] = "Sctp Host Id|2|1"  # NombreParametro|Posiciondel Site|Posicion del caberzera.
                        list_key_parameter_1Value[
                            "USERPLANEHOST"] = "User Plane Host Id|2|1"  # NombreParametro|Posiciondel Site|Posicion del caberzera.
                        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
                        Node_DF4G = ""

                        # validar que exista [NE NAME or ENODENAME]
                        Position_CellId_NE_ENODE_NAME1 = get_position_new_parameter(["Name"], Excel_sheet_CONF_SITE3)
                        position_id_cellId_NE_ENODE_NAME1 = Position_CellId_NE_ENODE_NAME1["Col_1"]

                        Position_CellId_NE_ENODE_NAME2 = get_position_new_parameter(["eNodeB Name"],
                                                                                    Excel_sheet_CONF_SITE3)
                        position_id_cellId_NE_ENODE_NAME2 = Position_CellId_NE_ENODE_NAME2["Col_1"]

                        if position_id_cellId_NE_ENODE_NAME1 == 1 and position_id_cellId_NE_ENODE_NAME2 == -1:
                            Node_DF4G = item_row_df4g["site_1"]["NE_Name_df4g"]
                        else:
                            Node_DF4G = item_row_df4g["site_1"]["eNodeB_Name_df4g"]
                        if position_id_cellId == -1:
                            if name_mo_df in list_evaluate1Value:
                                key_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[0]
                                position_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[1]
                                position_row = str(list_key_parameter_1Value[name_mo_df]).split('|')[2]
                                val = list_parameter[key_value + position_row]
                                for m in range(3, Excel_sheet_CONF_SITE3.max_row + 1):
                                    CS1_TYPE1_node = str(
                                        Excel_sheet_CONF_SITE3.cell(column=1,
                                                                    row=m).value).strip()
                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                        Excel_sheet_CONF_SITE3.cell(column=int(position_value),
                                                                    row=m).value).strip()
                                    if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                        if CS1_TYPE1_node.upper().strip() == Node_DF4G.upper().strip():
                                            if val == CS1_TYPE1_SEARCH_ROW_VALUE:
                                                key = key + 1
                                                col = 0
                                                list_parameter_value_site = {}
                                                for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                    col = col + 1
                                                    col_position = position_parameter["Col_" + str(col)]
                                                    if col_position > 0:
                                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                            Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                        row=m).value).strip()
                                                        list_parameter_value_site[
                                                            list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                    else:
                                                        list_parameter_value_site[
                                                            list_key_parameter] = 'NONE_COLUMN_FOUND'
                                                list_parameter_value_site_dic[
                                                    "key_" + str(key)] = list_parameter_value_site

                            else:
                                list_parameter_value_site_dic = {}
                                key = 0
                                for m in range(3, Excel_sheet_CONF_SITE3.max_row + 1):
                                    CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                        Excel_sheet_CONF_SITE3.cell(column=1,
                                                                    row=m).value).strip()
                                    if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                        # and CS1_TYPE1_SEARCH_ROW_VALUE =="":
                                        if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == Node_DF4G.upper().strip():
                                            key = key + 1
                                            col = 0

                                            list_parameter_value_site = {}
                                            for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                col = col + 1
                                                col_position = position_parameter["Col_" + str(col)]
                                                if col_position > 0:
                                                    CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                        Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                    row=m).value).strip()
                                                    list_parameter_value_site[
                                                        list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                    # CellAlgoSwitch
                                                else:
                                                    list_parameter_value_site[list_key_parameter] = 'NONE_COLUMN_FOUND'

                                    list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                            index_result_2 = self.write_ws2_sheet_others_value_correct_current(Cell_id_DF4G=0,
                                                                                               Enode=Node_DF4G,
                                                                                               list_parameter_mo=list_parameter,
                                                                                               list_parameter_site=list_parameter_value_site_dic,
                                                                                               index=index_result_2,
                                                                                               mo_name=name_mo_df)
                        else:
                            Cell_id_DF4G = ""
                            for item_row_parameter, item_row_parameter_value in item_row_df4g.items():
                                Cell_id_DF4G = item_row_parameter_value["CellId_df4g"]
                                list_parameter["Local Cell Id" + list_parameter_index] = Cell_id_DF4G
                                if name_mo_df in list_evaluate1Value:
                                    key_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[0]
                                    position_value = str(list_key_parameter_1Value[name_mo_df]).split('|')[1]
                                    position_row = str(list_key_parameter_1Value[name_mo_df]).split('|')[2]
                                    val = list_parameter[key_value + position_row]
                                    for m in range(3, Excel_sheet_CONF_SITE3.max_row + 1):
                                        # position_id_cellId
                                        Index_cell_id = str(
                                            Excel_sheet_CONF_SITE3.cell(column=int(position_id_cellId),
                                                                        row=m).value).strip()
                                        CS1_TYPE1_node = str(
                                            Excel_sheet_CONF_SITE3.cell(column=1,
                                                                        row=m).value).strip()
                                        CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                            Excel_sheet_CONF_SITE3.cell(column=int(position_value),
                                                                        row=m).value).strip()
                                        if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                            if CS1_TYPE1_node.upper().strip() == Node_DF4G.upper().strip() and Cell_id_DF4G == Index_cell_id:
                                                if val == CS1_TYPE1_SEARCH_ROW_VALUE:
                                                    key = key + 1
                                                    col = 0
                                                    list_parameter_value_site = {}
                                                    for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                        col = col + 1
                                                        col_position = position_parameter["Col_" + str(col)]
                                                        if col_position > 0:
                                                            CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                                Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                            row=m).value).strip()
                                                            list_parameter_value_site[
                                                                list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                        else:
                                                            list_parameter_value_site[
                                                                list_key_parameter] = 'NONE_COLUMN_FOUND'
                                                    list_parameter_value_site_dic[
                                                        "key_" + str(key)] = list_parameter_value_site

                                else:
                                    list_parameter_value_site_dic = {}
                                    key = 0
                                    for m in range(3, Excel_sheet_CONF_SITE3.max_row + 1):
                                        CS1_TYPE1_SEARCH_ROW_VALUE = ""
                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                            Excel_sheet_CONF_SITE3.cell(column=1,
                                                                        row=m).value).strip()
                                        Index_cell_id = str(
                                            Excel_sheet_CONF_SITE3.cell(column=int(position_id_cellId),
                                                                        row=m).value).strip()
                                        if len(CS1_TYPE1_SEARCH_ROW_VALUE) > 0 and CS1_TYPE1_SEARCH_ROW_VALUE != 'None':
                                            # and CS1_TYPE1_SEARCH_ROW_VALUE =="":
                                            if CS1_TYPE1_SEARCH_ROW_VALUE.upper().strip() == Node_DF4G.upper().strip() and Cell_id_DF4G == Index_cell_id:
                                                key = key + 1
                                                col = 0

                                                list_parameter_value_site = {}
                                                for list_key_parameter, list_key_parameter_value in list_parameter.items():
                                                    col = col + 1
                                                    col_position = position_parameter["Col_" + str(col)]
                                                    if col_position > 0:
                                                        CS1_TYPE1_SEARCH_ROW_VALUE = str(
                                                            Excel_sheet_CONF_SITE3.cell(column=col_position,
                                                                                        row=m).value).strip()
                                                        list_parameter_value_site[
                                                            list_key_parameter] = CS1_TYPE1_SEARCH_ROW_VALUE
                                                        # CellAlgoSwitch
                                                    else:
                                                        list_parameter_value_site[
                                                            list_key_parameter] = 'NONE_COLUMN_FOUND'

                                        list_parameter_value_site_dic["key_" + str(key)] = list_parameter_value_site
                                index_result_2 = self.write_ws2_sheet_others_value_correct_current(
                                    Cell_id_DF4G=int(Cell_id_DF4G), Enode=Node_DF4G,
                                    list_parameter_mo=list_parameter,
                                    list_parameter_site=list_parameter_value_site_dic,
                                    index=index_result_2,
                                    mo_name=name_mo_df)
        list_output["index_result_1"] = index_result_1
        list_output["index_result_2"] = index_result_2
        return list_output

    def validate_sheet_UCELLNFREQPRIOINFO(self, p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet,
                                          index_result_2: int):
        list_parameter_df: list = ["Cell Name",
                                   "Logical RNC ID", "Cell ID", "E-UTRA Absolute Radio Frequency Channel Number",
                                   "Absolute Priority Level of the E-UTRA Frequency",
                                   "Measurement Bandwidth"]

        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        flat_row = False
        row_identificated = ""
        # recorrer cada registro.
        for i in range(2, p_excel_sheet_df.max_row + 1):
            RNC_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
            if len(RNC_Name_df) > 0 and RNC_Name_df != 'None' and RNC_Name_df != 'NONE':
                Cell_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                Logical_RNC_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                Cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                E_UTRA_Absolute_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                E_UTRA_Frequency_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                Measurement_Bandwidth_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                RAN_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                row_identificated = Cell_ID_df + '-' + RNC_Name_df
                if RAN_df == "MOVISTAR":
                    list_value_correct = []
                    # list_value_correct(RNC_Name_df)
                    list_value_correct.append(Cell_Name_df)
                    list_value_correct.append(Logical_RNC_ID_df)
                    list_value_correct.append(Cell_ID_df)
                    list_value_correct.append(E_UTRA_Absolute_df)
                    list_value_correct.append(E_UTRA_Frequency_df)
                    list_value_correct.append(Measurement_Bandwidth_df)
                    self.write_ws2_result_mo_parameter(row_identificated, "UCELLNFREQPRIOINFO",
                                                       list_parameter_df,
                                                       index_result_2)
                    for u in range(3, excel_sheet_site.max_row + 1):
                        Cell_Name_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                        Cell_RNC_ID_site = str(excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                        Cell_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                        if Cell_Name_df == Cell_Name_site and Logical_RNC_ID_df == Cell_RNC_ID_site \
                                and Cell_ID_df == Cell_ID_site:

                            E_UTRA_Absolute_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                            E_UTRA_Frequency_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                            Measurement_Bandwidth_site = str(
                                excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                            list_value_current = []
                            list_value_current.append(Cell_Name_site)
                            list_value_current.append(Cell_RNC_ID_site)
                            list_value_current.append(Cell_ID_site)
                            list_value_current.append(E_UTRA_Absolute_site)
                            list_value_current.append(E_UTRA_Frequency_site)
                            list_value_current.append(Measurement_Bandwidth_site)
                            flat_row = True

                            index_result_2 = self.write_ws2_result_value_correct_current(list_parameter_mo={},
                                                                                         list_parameter_df=list_parameter_df,
                                                                                         list_value_correct=list_value_correct,
                                                                                         list_value_current=list_value_current,
                                                                                         index=index_result_2,
                                                                                         flat_fail=False,
                                                                                         message="")
                            break
                        else:
                            flat_row = False
                    if flat_row is False:
                        index_result_2 = self.write_ws2_result_value_correct_current(list_parameter_mo={},
                                                                                     list_parameter_df=list_parameter_df,
                                                                                     list_value_correct=list_value_correct,
                                                                                     list_value_current=list_value_current,
                                                                                     index=index_result_2,
                                                                                     flat_fail=True,
                                                                                     message="Config Site empty")
        return index_result_2

    def validate_sheet_ULTENCELL(self, p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int):
        list_parameter_df: list = ["RNC ID", "Cell ID",
                                   "Cell Name", "LTE Cell Index", "LTE Cell Name",
                                   "Blind HO Neighboring Cell Flag"]

        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        flat_row = False
        row_identificated = ""
        # recorrer cada registro.
        for i in range(2, p_excel_sheet_df.max_row + 1):
            RNC_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
            if len(RNC_Name_df) > 0 and RNC_Name_df != 'None' and RNC_Name_df != 'NONE':
                RNC_ID_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                Cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                Cell_Name_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                LTE_Cell_Index_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                LTE_Cell_Name_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                Blind_HO_Neighboring_Cell_Flag_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                RAN_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()

                row_identificated = Cell_ID_df + '-' + RNC_Name_df

                if RAN_df == "MOVISTAR":
                    list_value_correct = []
                    # list_value_correct(RNC_Name_df)
                    list_value_correct.append(RNC_ID_df)
                    list_value_correct.append(Cell_ID_df)
                    list_value_correct.append(Cell_Name_df)
                    list_value_correct.append(LTE_Cell_Index_df)
                    list_value_correct.append(LTE_Cell_Name_df)
                    list_value_correct.append(Blind_HO_Neighboring_Cell_Flag_df)

                    for u in range(3, excel_sheet_site.max_row + 1):
                        RNC_ID_site = str(excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                        Cell_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                        Cell_Name_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                        if Cell_Name_df == Cell_Name_site and RNC_ID_site == RNC_ID_df \
                                and Cell_ID_df == Cell_ID_site:
                            self.write_ws2_result_mo_parameter(row_identificated, "ULTENCELL",
                                                               list_parameter_df,
                                                               index_result_2)

                            LTE_Cell_Index_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                            LTE_Cell_Name_site = str(excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                            Blind_HO_Neighboring_Cell_Flag_site = str(
                                excel_sheet_site.cell(row=u, column=8).value).strip().upper()

                            list_value_current = []
                            list_value_current.append(RNC_ID_site)
                            list_value_current.append(Cell_ID_site)
                            list_value_current.append(Cell_Name_site)
                            list_value_current.append(LTE_Cell_Index_site)
                            list_value_current.append(LTE_Cell_Name_site)
                            list_value_current.append(Blind_HO_Neighboring_Cell_Flag_site)
                            flat_row = True

                            index_result_2 = self.write_ws2_result_value_correct_current(list_parameter_mo={},
                                                                                         list_parameter_df=list_parameter_df,
                                                                                         list_value_correct=list_value_correct,
                                                                                         list_value_current=list_value_current,
                                                                                         index=index_result_2,
                                                                                         flat_fail=False,
                                                                                         message="")
                            break
                        else:
                            flat_row = False
                    if flat_row is False:
                        index_result_2 = self.write_ws2_result_value_correct_current(list_parameter_mo={},
                                                                                     list_parameter_df=list_parameter_df,
                                                                                     list_value_correct=list_value_correct,
                                                                                     list_value_current=list_value_current,
                                                                                     index=index_result_2,
                                                                                     flat_fail=True,
                                                                                     message="Config Site empty")
        return index_result_2

    # validar external
    def validate_utranncell_external(self,
                                     p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                     list_parameter_mo: dict,
                                     NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        Cell_id_DF4G = ""
        row_identificated = ""
        list_parameter_df: list = ["Local Cell ID", "Rnc Id", "Rnc Cell Id", "Blind Handover Priority",
                                   "Local Cell Name",
                                   "Neighbour Cell Name", "Mobile Country Code", "Mobile Network Code"]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-UTRANNCELL-" + NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and value_Ne_Name_df == NE_NAME_df_filter:

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    RNC_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    RNC_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Blind_handover_priority_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Local_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Neighbour_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    list_value_correct = []

                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(RNC_ID_df)
                    list_value_correct.append(RNC_cell_ID_df)
                    list_value_correct.append(Blind_handover_priority_df)
                    list_value_correct.append(Local_cell_name_df)
                    list_value_correct.append(Neighbour_cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Local_cell_ID_df == Cell_id_DF4G:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-UTRANNCELL", list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):
                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Neighbour_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=15).value).strip().upper()
                                Local_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=14).value).strip().upper()

                                Local_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                            if Neighbour_cell_name_site == Neighbour_cell_name_df \
                                    and Local_cell_name_site == Local_cell_name_df and Local_cell_ID_site == Local_cell_ID_df:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                RNC_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                RNC_cell_ID_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Blind_handover_priority_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                # Local_cell_name_site = str(excel_sheet_site.cell(row=u, column=14).value).strip().upper()
                                # Neighbour_cell_name_site=str(excel_sheet_site.cell(row=u, column=15).value).strip().upper
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(RNC_ID_site)
                                list_value_current.append(RNC_cell_ID_site)
                                list_value_current.append(Blind_handover_priority_site)
                                list_value_current.append(Local_cell_name_site)
                                list_value_current.append(Neighbour_cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_utranexternalcell_external(self,
                                            p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet,
                                            index_result_2: int, list_parameter_mo: dict,
                                            NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        row_identificated = ""
        list_parameter_df: list = ["Rnc Id", "Rnc Cell Id", "Downlink Uarfcn", "Uplink Uarfcn Configure Indicator",
                                   "Utran Cell Type Indicator", "Routing area code configure indicator",
                                   "Routing Area Code", "Primary Scrambling Code", "Location Area Code", "Cell Name",
                                   "Mobile Country Code", "Mobile Network Code"]
        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = [],
        list_value_current = []
        lis_value_message = []
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)
        ifv = 0
        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            ifv = ifv + 1
            if ifv > 1:
                break
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            row_identificated = NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and value_Ne_Name_df == NE_NAME_df_filter:
                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    # Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    RNC_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    RNC_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Downlink_UARFCN_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Uplink_UARFCN_configure_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    UTRAN_cell_type_indicator_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Routing_area_code_configure_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Routing_area_code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Primary_scrambling_code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    Location_area_code_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    Cell_name_df = str(p_excel_sheet_df.cell(row=i, column=13).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=14).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=15).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=16).value).strip().upper()
                    list_value_correct = []
                    row_identificated = Cell_name_df + "-" + RNC_cell_ID_df

                    # Agregar los valores configurados actual de Config DF
                    list_value_correct.append(RNC_ID_df)
                    list_value_correct.append(RNC_cell_ID_df)
                    list_value_correct.append(Downlink_UARFCN_df)
                    list_value_correct.append(Uplink_UARFCN_configure_indicator_df)
                    list_value_correct.append(UTRAN_cell_type_indicator_df)
                    list_value_correct.append(Routing_area_code_configure_indicator_df)
                    list_value_correct.append(Routing_area_code_df)
                    list_value_correct.append(Primary_scrambling_code_df)
                    list_value_correct.append(Location_area_code_df)
                    list_value_correct.append(Cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if "MOVISTAR" == RAN_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-UTRANEXTERNALCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):
                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]

                                cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=14).value).upper().strip()

                            if cell_name_site == Cell_name_df:

                                RNC_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                RNC_cell_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                Downlink_UARFCN_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Uplink_UARFCN_configure_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                UTRAN_cell_type_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Routing_area_code_configure_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                Routing_area_code_site = str(
                                    excel_sheet_site.cell(row=u, column=11).value).strip().upper()
                                Primary_scrambling_code_site = str(
                                    excel_sheet_site.cell(row=u, column=12).value).strip().upper()
                                Location_area_code_site = str(
                                    excel_sheet_site.cell(row=u, column=13).value).strip().upper()
                                Cell_name_site = str(excel_sheet_site.cell(row=u, column=14).value).strip().upper()
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()

                                list_value_current = []
                                # Agregar los valores configurados actual de Config site
                                list_value_current.append(RNC_ID_site)
                                list_value_current.append(RNC_cell_ID_site)
                                list_value_current.append(Downlink_UARFCN_site)
                                list_value_current.append(Uplink_UARFCN_configure_indicator_site)
                                list_value_current.append(UTRAN_cell_type_indicator_site)
                                list_value_current.append(Routing_area_code_configure_indicator_site)
                                list_value_current.append(Routing_area_code_site)
                                list_value_current.append(Primary_scrambling_code_site)
                                list_value_current.append(Location_area_code_site)
                                list_value_current.append(Cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_utrannfreq_external(self,
                                     p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet, index_result_2: int,
                                     list_parameter_mo: dict,
                                     NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "Downlink UARFCN", "Minimum required quality level",
                                   "Uplink UARFCN indicator",
                                   "Reselection priority configure indicator", "Cell reselection priority",
                                   "Frequency Priority for Connected Mode"]
        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = [],
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)
        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-UTRANNFREQ-" + NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and NE_NAME_df_filter == value_Ne_Name_df:
                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    Downlink_UARFCN_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Minimum_required_quality_level_df = str(
                        p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Uplink_UARFCN_indicator_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Reselection_priority_configure_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Cell_reselection_priority_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Frequency_Priority_for_Connected_Mode_df = str(
                        p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    list_value_correct = []

                    # Agregar los valores configurados actual de Config DF
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(Downlink_UARFCN_df)
                    list_value_correct.append(Minimum_required_quality_level_df)
                    list_value_correct.append(Uplink_UARFCN_indicator_df)
                    list_value_correct.append(Reselection_priority_configure_indicator_df)
                    list_value_correct.append(Cell_reselection_priority_df)
                    list_value_correct.append(Frequency_Priority_for_Connected_Mode_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-UTRANNFREQ", list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):
                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).upper().strip()
                                Downlink_UARFCN_site = str(excel_sheet_site.cell(row=u, column=3).value).upper().strip()

                            if Local_cell_ID_site == Local_cell_ID_df and Downlink_UARFCN_site == Downlink_UARFCN_df:

                                RNC_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                list_value_current = []
                                # Agregar los valores configurados actual de Config site

                                Minimum_required_quality_level_site = str(
                                    excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Uplink_UARFCN_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                Reselection_priority_configure_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Cell_reselection_priority_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                Frequency_Priority_for_Connected_Mode_site = str(
                                    excel_sheet_site.cell(row=u, column=20).value).strip().upper()

                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(Downlink_UARFCN_site)
                                list_value_current.append(Minimum_required_quality_level_site)
                                list_value_current.append(Uplink_UARFCN_indicator_site)
                                list_value_current.append(Reselection_priority_configure_indicator_site)
                                list_value_current.append(Cell_reselection_priority_site)
                                list_value_current.append(Frequency_Priority_for_Connected_Mode_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_eutraninterfreqncell_external(self,
                                               p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet,
                                               index_result_2: int, list_parameter_mo: dict,
                                               NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "eNodeB ID", "NCell ID", "Blind handover Priority",
                                   "Local cell name",
                                   "Neighbour Cell Name", "Mobile Country Code", "Mobile Network Code"]
        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANINTERFREQNCELL-" + NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and NE_NAME_df_filter == value_Ne_Name_df:

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    eNodeB_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    NCell_ID = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Blind_handover_priority_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Local_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Neighbour_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(eNodeB_ID_df)
                    list_value_correct.append(NCell_ID)
                    list_value_correct.append(Blind_handover_priority_df)
                    list_value_correct.append(Local_cell_name_df)
                    list_value_correct.append(Neighbour_cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANINTERFREQNCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Neighbour_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=14).value).strip().upper()
                                Local_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=13).value).strip().upper()

                                Local_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                            if Neighbour_cell_name_site == Neighbour_cell_name_df \
                                    and Local_cell_name_site == Local_cell_name_df and Local_cell_ID_site == Local_cell_ID_df:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                eNodeB_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                NCell_ID = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Blind_handover_priority_site = str(
                                    excel_sheet_site.cell(row=u, column=11).value).strip().upper()
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(eNodeB_ID_site)
                                list_value_current.append(NCell_ID)
                                list_value_current.append(Blind_handover_priority_site)
                                list_value_current.append(Local_cell_name_site)
                                list_value_current.append(Neighbour_cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_eutraninternfreq_external(self,
                                           p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet,
                                           index_result_2: int, list_parameter_mo: dict,
                                           NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "Downlink EARFCN", "Uplink EARFCN indicator",
                                   "Inter Frequency cell resel priority indicator",
                                   "Inter Frequency cell resel priority",
                                   "EUTRAN reselection time", "Speed dependent resel parameter configuring indicator",
                                   "Measurement bandwidth",
                                   "P Max configuring indicator"]

        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANINTERNFREQ-" + NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and NE_NAME_df_filter == value_Ne_Name_df:

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    Downlink_EARFCN_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Uplink_EARFCN_indicator_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Inter_Frequency_cell_resel_priority_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Inter_Frequency_cell_resel_priority_df = str(
                        p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    EUTRAN_reselection_time_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Speed_dependent_resel_parameter_configuring_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Measurement_bandwidth_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    P_Max_configuring_indicator_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    Q_Qual_Min_configuring_indicator_df = str(
                        p_excel_sheet_df.cell(row=i, column=13).value).strip().upper()
                    Clasificacion_df = str(p_excel_sheet_df.cell(row=i, column=14).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=15).value).strip().upper()

                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(Downlink_EARFCN_df)
                    list_value_correct.append(Uplink_EARFCN_indicator_df)
                    list_value_correct.append(Inter_Frequency_cell_resel_priority_indicator_df)
                    list_value_correct.append(Inter_Frequency_cell_resel_priority_df)
                    list_value_correct.append(EUTRAN_reselection_time_df)
                    list_value_correct.append(Speed_dependent_resel_parameter_configuring_indicator_df)
                    list_value_correct.append(Measurement_bandwidth_df)
                    list_value_correct.append(P_Max_configuring_indicator_df)
                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANINTERNFREQ",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Local_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()

                                Downlink_EARFCN_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                            if  Local_cell_ID_df == Local_cell_ID_site and Downlink_EARFCN_df == Downlink_EARFCN_site:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                Downlink_EARFCN_site = str(excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Uplink_EARFCN_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                Inter_Frequency_cell_resel_priority_indicator_site = \
                                    str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Inter_Frequency_cell_resel_priority_site = \
                                    str(excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                EUTRAN_reselection_time_site = str(
                                    excel_sheet_site.cell(row=u, column=8).value).strip().upper()
                                Speed_dependent_resel_parameter_configuring_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Measurement_bandwidth_site = str(
                                    excel_sheet_site.cell(row=u, column=12).value).strip().upper()
                                P_Max_configuring_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=17).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(Downlink_EARFCN_site)
                                list_value_current.append(Uplink_EARFCN_indicator_site)
                                list_value_current.append(Inter_Frequency_cell_resel_priority_indicator_site)
                                list_value_current.append(Inter_Frequency_cell_resel_priority_site)
                                list_value_current.append(EUTRAN_reselection_time_site)
                                list_value_current.append(Speed_dependent_resel_parameter_configuring_indicator_site)
                                list_value_current.append(Measurement_bandwidth_site)
                                list_value_current.append(P_Max_configuring_indicator_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current, sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position, position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
                    # else:
                    # index_result_2 = index_result_2 + 1
                    # write_ws2_result_mo_parameter("VECINDAD-EUTRANINTERNFREQ", list_parameter_df, index_result_2)
                    # index_result_2 = write_ws2_result_value_correct_current(list_parameter_mo=list_parameter_mo,
                    #                                                        list_parameter_df=list_parameter_df,
                    #                                                        list_value_correct=list_value_correct,
                    #                                                        list_value_current=list_value_current,
                    #                                                        index=index_result_2,
                    #                                                        flat_fail=True,
                    #                                                        message="The [NE Name] :{} or "
                    #                                                                "[eNodeB Name] :{} incorrect"
                    #                                                        .format(value_Ne_Name_df,
                    #                                                                value_eNodeB_Name_df))
        return index_result_2

    def validate_eutranintrafreqncell_external(self,
                                               p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet,
                                               index_result_2: int, list_parameter_mo: dict,
                                               NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Local cell ID", "eNodeB ID", "Cell ID",
                                   "Local cell name",
                                   "Neighbour cell name",
                                   "Mobile Country Code", "Mobile Network Code"]

        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        key_row = ""
        value_row = ""
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANINTRAFREQNCELL-" + NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and NE_NAME_df_filter == value_Ne_Name_df:

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Local_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    eNodeB_ID_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    Cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Local_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Neighbour_cell_name_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Mobile_Country_Code_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Mobile_Network_Code_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()

                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Local_cell_ID_df)
                    list_value_correct.append(eNodeB_ID_df)
                    list_value_correct.append(Cell_ID_df)
                    list_value_correct.append(Local_cell_name_df)
                    list_value_correct.append(Neighbour_cell_name_df)
                    list_value_correct.append(Mobile_Country_Code_df)
                    list_value_correct.append(Mobile_Network_Code_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Cell_id_DF4G == Local_cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANINTRAFREQNCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Local_ell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=12).value).strip().upper()

                                Neighbour_cell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=13).value).strip().upper()
                            if Local_cell_name_df == Local_ell_name_site and Neighbour_cell_name_df == Neighbour_cell_name_site:

                                Local_cell_ID_site = str(excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                eNodeB_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                Cell_ID_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Mobile_Country_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                Mobile_Network_Code_site = str(
                                    excel_sheet_site.cell(row=u, column=4).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Local_cell_ID_site)
                                list_value_current.append(eNodeB_ID_site)
                                list_value_current.append(Cell_ID_site)
                                list_value_current.append(Local_ell_name_site)
                                list_value_current.append(Neighbour_cell_name_site)
                                list_value_current.append(Mobile_Country_Code_site)
                                list_value_current.append(Mobile_Network_Code_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current,
                                    sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position,
                                    position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_eutranexternalcell_external(self,
                                             p_excel_sheet_df: Worksheet, excel_sheet_site: Worksheet,
                                             index_result_2: int, list_parameter_mo: dict,
                                             NE_NAME_df_filter: str):
        item_row_df4g = get_df4g_row_site(self.excel_document_DF["DF 4G"])
        value_Ne_Name = ""
        value_eNodeB_Name = ""
        Ne_Name_DF4G = ""
        eNodeB_Name_DF4G = ""
        list_parameter_df: list = ["Mobile country code", "Mobile network code",
                                   "eNodeB ID", "Cell ID", "Downlink EARFCN", "Uplink EARFCN indicator",
                                   "Physical cell ID", "Tracking area code", "Cell name"]

        list_parameter_df = [element.title() for element in list_parameter_df]
        list_value_correct = []
        list_value_current = []
        lis_value_message = []
        row_identificated = ""
        Cell_id_DF4G = ""
        list_new_parameter = get_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                     list_parameter_lte=list_parameter_mo)
        list_new_parameter_position = get_position_new_parameter(list_new_parameter=list_new_parameter,
                                                                 sheet_name_site=excel_sheet_site)

        list_parameter_df = add_new_parameter_by_mo(list_parameter_df=list_parameter_df,
                                                    list_new_parameter=list_new_parameter)

        # Recorremos las filas de DF4G
        for key_row, value_row in item_row_df4g.items():
            Ne_Name_DF4G = value_row["NE_Name_df4g"]
            eNodeB_Name_DF4G = value_row["eNodeB_Name_df4g"]
            Cell_id_DF4G = value_row["CellId_df4g"]
            row_identificated = str(Cell_id_DF4G) + "-EUTRANEXTERNALCELL-" + NE_NAME_df_filter
            flat_row = False  # Valida que exista un valor con los parametros de filtro en el caso no se encuentre todos los parametros serán invalidos.
            # recorremos los  valores de DF
            # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
            for i in range(2, p_excel_sheet_df.max_row + 1):
                value_Ne_Name_df = str(p_excel_sheet_df.cell(row=i, column=2).value).strip().upper()
                value_eNodeB_Name_df = str(p_excel_sheet_df.cell(row=i, column=3).value).strip().upper()
                if len(
                        value_Ne_Name_df) > 0 and value_Ne_Name_df != 'None' and value_Ne_Name_df != 'NONE' and NE_NAME_df_filter == value_Ne_Name_df:

                    # Recorrer solo los valores de NE_Name :*eNodeB Name  que se encuentra en la hoja DF4G
                    # Obtenemos el valor que se encuentra en el DF
                    Mobile_country_code_df = str(p_excel_sheet_df.cell(row=i, column=4).value).strip().upper()
                    Mobile_network_code_df = str(p_excel_sheet_df.cell(row=i, column=5).value).strip().upper()
                    eNodeB_ID_df = str(p_excel_sheet_df.cell(row=i, column=6).value).strip().upper()
                    Cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=7).value).strip().upper()
                    Downlink_EARFCN_df = str(p_excel_sheet_df.cell(row=i, column=8).value).strip().upper()
                    Uplink_EARFCN_indicator_df = str(p_excel_sheet_df.cell(row=i, column=9).value).strip().upper()
                    Physical_cell_ID_df = str(p_excel_sheet_df.cell(row=i, column=10).value).strip().upper()
                    Tracking_area_code_df = str(p_excel_sheet_df.cell(row=i, column=11).value).strip().upper()
                    Cell_name_df = str(p_excel_sheet_df.cell(row=i, column=12).value).strip().upper()
                    RAN_df = str(p_excel_sheet_df.cell(row=i, column=13).value).strip().upper()

                    list_value_correct = []
                    # Agregar los valores configurados actual de Config
                    list_value_correct.append(Mobile_country_code_df)
                    list_value_correct.append(Mobile_network_code_df)
                    list_value_correct.append(eNodeB_ID_df)
                    list_value_correct.append(Cell_ID_df)
                    list_value_correct.append(Downlink_EARFCN_df)
                    list_value_correct.append(Uplink_EARFCN_indicator_df)
                    list_value_correct.append(Physical_cell_ID_df)
                    list_value_correct.append(Tracking_area_code_df)
                    list_value_correct.append(Cell_name_df)

                    list_value_correct = add_new_parameter_correct_value(list_value_output=list_value_correct,
                                                                         list_parameter_lte=list_parameter_mo,
                                                                         list_new_parameter=list_new_parameter)

                    if RAN_df == "MOVISTAR" and Cell_id_DF4G == Cell_ID_df:
                        # se debe de buscar_df las filas en el config del site  con los filtros
                        # Filter: [eNodeB Name,Neighbour cell name,Local cell name,Local cell ID]
                        # Recorrer Site
                        # index_result_2 = index_result_2 + 1
                        self.write_ws2_result_mo_parameter(row_identificated, "VECINDAD-EUTRANEXTERNALCELL",
                                                           list_parameter_df,
                                                           index_result_2)
                        for u in range(3, excel_sheet_site.max_row + 1):

                            eNodeB_Name_site = str(excel_sheet_site.cell(row=u, column=1).value).strip().upper()
                            # Validamos que la información del site tenga información.
                            if len(eNodeB_Name_site) > 0 and eNodeB_Name_site != 'None' and eNodeB_Name_site != 'NONE':
                                # Filtramos la información Filter: [eNodeB Name,Neighbour cell name,
                                # Local cell name,Local cell ID]
                                Local_ell_name_site = str(
                                    excel_sheet_site.cell(row=u, column=11).value).strip().upper()

                            if Cell_name_df == Local_ell_name_site:

                                Mobile_country_code_site = str(
                                    excel_sheet_site.cell(row=u, column=2).value).strip().upper()
                                Mobile_network_code_site = str(
                                    excel_sheet_site.cell(row=u, column=3).value).strip().upper()
                                eNodeB_ID_site = str(excel_sheet_site.cell(row=u, column=4).value).strip().upper()
                                Cell_ID_site = str(excel_sheet_site.cell(row=u, column=5).value).strip().upper()
                                Downlink_EARFCN_site = str(excel_sheet_site.cell(row=u, column=6).value).strip().upper()
                                Uplink_EARFCN_indicator_site = str(
                                    excel_sheet_site.cell(row=u, column=7).value).strip().upper()
                                Physical_cell_ID_site = str(
                                    excel_sheet_site.cell(row=u, column=9).value).strip().upper()
                                Tracking_area_code_site = str(
                                    excel_sheet_site.cell(row=u, column=10).value).strip().upper()
                                Cell_name_site = str(excel_sheet_site.cell(row=u, column=11).value).strip().upper()

                                list_value_current = []
                                list_value_current.append(Mobile_country_code_site)
                                list_value_current.append(Mobile_network_code_site)
                                list_value_current.append(eNodeB_ID_site)
                                list_value_current.append(Cell_ID_site)
                                list_value_current.append(Downlink_EARFCN_site)
                                list_value_current.append(Uplink_EARFCN_indicator_site)
                                list_value_current.append(Physical_cell_ID_site)
                                list_value_current.append(Tracking_area_code_site)
                                list_value_current.append(Cell_name_site)

                                list_value_current = add_new_parameter_current_value(
                                    list_value_output=list_value_current,
                                    sheet_name_site=excel_sheet_site,
                                    list_new_parameter_position=list_new_parameter_position,
                                    position_row=u)

                                index_result_2 = self.write_ws2_result_value_correct_current(
                                    list_parameter_mo=list_parameter_mo,
                                    list_parameter_df=list_parameter_df,
                                    list_value_correct=list_value_correct,
                                    list_value_current=list_value_current,
                                    index=index_result_2,
                                    flat_fail=False,
                                    message="")
                                flat_row = True
                                break
                            else:
                                flat_row = False
                        if flat_row is False:
                            index_result_2 = self.write_ws2_result_value_correct_current(
                                list_parameter_mo=list_parameter_mo,
                                list_parameter_df=list_parameter_df,
                                list_value_correct=list_value_correct,
                                list_value_current=list_value_current,
                                index=index_result_2,
                                flat_fail=True,
                                message="Config Site empty")
        return index_result_2

    def validate_external_huawei(self, name: str, workbook_site1: Workbook, index_result_1: int, index_result_2: int):
        list_parameter = {}
        list_output = {}
        str_prc = name
        position_4g = str_prc.find('4G')
        if position_4g > 0:
            NE_NAME_df_filter = str(str_prc[position_4g + 3:]).upper()
            DF_array = str_prc[:position_4g - 1].split('_')
            # Recorremos todos los external
            for name_mo_df in DF_array:
                name_mo_df = name_mo_df.upper().strip()

                Excel_sheet_DF = self.excel_document_DF[name_mo_df]
                excel_sheet_lte = self.excel_document_LTE[self.excel_NameSheet_LTE]
                list_parameter = get_parameter_by_mo(excel_sheet_lte, str(name_mo_df).upper())
                # validar que exista la hoja [MO] en Confidata si no existe enviar el RnpData_BTS3900
                index_result_1 = index_result_1 + 1

                list_sheet_namessite = list(workbook_site1.sheetnames)

                list_sheet_namessite_Upper = [element.upper() for element in list_sheet_namessite]

                if name_mo_df in list_sheet_namessite_Upper:
                    position_lista = list_sheet_namessite_Upper.index(name_mo_df)
                    nuevanombreHoja = list_sheet_namessite[position_lista]
                    Excel_sheet_CONF_SITE1 = workbook_site1[nuevanombreHoja]
                    # Validamos cada uno de las hojas:
                    if name_mo_df == "UTRANNCELL":
                        index_result_2 = self.validate_utranncell_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                           excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                           index_result_2=index_result_2,
                                                                           list_parameter_mo=list_parameter,
                                                                           NE_NAME_df_filter=NE_NAME_df_filter)
                    if name_mo_df == "UTRANEXTERNALCELL":
                        index_result_2 = self.validate_utranexternalcell_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                                  excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                                  index_result_2=index_result_2,
                                                                                  list_parameter_mo=list_parameter,
                                                                                  NE_NAME_df_filter=NE_NAME_df_filter)
                    if name_mo_df == "UTRANNFREQ":
                        index_result_2 = self.validate_utrannfreq_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                           excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                           index_result_2=index_result_2,
                                                                           list_parameter_mo=list_parameter,
                                                                           NE_NAME_df_filter=NE_NAME_df_filter)
                    if name_mo_df == "EUTRANINTERFREQNCELL":
                        index_result_2 = self.validate_eutraninterfreqncell_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                                     excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                                     index_result_2=index_result_2,
                                                                                     list_parameter_mo=list_parameter,
                                                                                     NE_NAME_df_filter=NE_NAME_df_filter)
                    if name_mo_df == "EUTRANINTERNFREQ":
                        index_result_2 = self.validate_eutraninternfreq_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                                 excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                                 index_result_2=index_result_2,
                                                                                 list_parameter_mo=list_parameter,
                                                                                 NE_NAME_df_filter=NE_NAME_df_filter)
                    if name_mo_df == "EUTRANINTRAFREQNCELL":
                        index_result_2 = self.validate_eutranintrafreqncell_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                                     excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                                     index_result_2=index_result_2,
                                                                                     list_parameter_mo=list_parameter,
                                                                                     NE_NAME_df_filter=NE_NAME_df_filter)
                    if name_mo_df == "EUTRANEXTERNALCELL":
                        index_result_2 = self.validate_eutranexternalcell_external(p_excel_sheet_df=Excel_sheet_DF,
                                                                                   excel_sheet_site=Excel_sheet_CONF_SITE1,
                                                                                   index_result_2=index_result_2,
                                                                                   list_parameter_mo=list_parameter,
                                                                                   NE_NAME_df_filter=NE_NAME_df_filter)

        list_output["index_result_1"] = index_result_1
        list_output["index_result_2"] = index_result_2
        return list_output

    def validate_external_Huawei_final(self, sitename: str, url_site: str, index_result_1: int, index_result_2: int):
        list_output = {}

        excel_document_site = openpyxl.load_workbook(filename=url_site)

        list_output = self.validate_external_huawei(sitename, excel_document_site, index_result_1, index_result_2)

        index_result_1 = list_output["index_result_1"]
        index_result_2 = list_output["index_result_2"]

        list_output["index_result_1"] = index_result_1
        list_output["index_result_2"] = index_result_2
        return list_output
